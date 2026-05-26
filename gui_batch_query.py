import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import requests
import json
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
import csv
import time
import os
import threading
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
import shutil

# 接口地址
API_URL = "https://b.cainiao.com/merchantchargeorder/queryChargeOrderListNew"

# 默认Cookie (留空，让用户自己填)
DEFAULT_COOKIE = ""

def get_charge_order_list(order_no, cookie):
    """
    发送POST请求获取菜鸟充值订单列表
    """
    # 模拟浏览器的请求头
    headers = {
        "accept": "application/json, text/plain, */*",
        "bx-v": "2.5.36",
        "content-type": "application/json",
        "h-csrf": "635e77ca-003a-498a-9a58-18c213306b08",
        "referer": "https://b.cainiao.com/cf_seller/charge-order/charge-order-query",
        "sec-ch-ua": '"Not:A-Brand";v="99", "Microsoft Edge";v="145", "Chromium";v="145"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36 Edg/145.0.0.0",
        "Cookie": cookie
    }

    # POST请求参数
    post_data = {
        "orderNo": order_no,
        "currentPage": 1,
        "pageSize": 10,
        "orderNoType": 0
    }

    try:
        # 发送POST请求
        response = requests.post(
            url=API_URL,
            headers=headers,
            data=json.dumps(post_data),
            timeout=30,
            allow_redirects=False
        )

        # 检查是否发生重定向
        if response.status_code in (301, 302, 303, 307, 308):
            return {"error": f"请求被重定向 (Cookie可能失效)"}

        # 检查响应状态码
        if response.status_code != 200:
            return {"error": f"请求失败: 状态码 {response.status_code}"}

        return response.json()

    except Exception as e:
        return {"error": f"请求异常: {e}"}

class BatchQueryApp:
    def __init__(self, root):
        self.root = root
        self.root.title("菜鸟充值订单批量查询工具 v1.0")
        self.root.geometry("800x600")
        
        # 变量
        self.file_path = tk.StringVar()
        self.file_paths = []
        self.cookie_var = tk.StringVar(value=DEFAULT_COOKIE)
        self.exclude_product_ids_var = tk.StringVar()
        self.status_var = tk.StringVar(value="准备就绪")
        self.progress_var = tk.DoubleVar()
        self.thread_count = tk.IntVar(value=6)
        self.log_messages = []
        self.pending_files = []
        self.last_output_file = None
        self.total_files = 0
        self.processed_files = 0
        
        self.config_file = "cookie.txt"
        self.exclude_ids_file = "exclude_ids.txt"
        self.load_config()
        
        self.create_widgets()
        
    def load_config(self):
        # 加载 Cookie
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, "r", encoding="utf-8") as f:
                    cookie = f.read().strip()
                    if cookie:
                        self.cookie_var.set(cookie)
            except:
                pass
                
        # 加载排除商品ID
        if os.path.exists(self.exclude_ids_file):
            try:
                with open(self.exclude_ids_file, "r", encoding="utf-8") as f:
                    exclude_ids = f.read().strip()
                    if exclude_ids:
                        self.exclude_product_ids_var.set(exclude_ids)
            except:
                pass

    def save_config(self):
        # 保存 Cookie
        try:
            with open(self.config_file, "w", encoding="utf-8") as f:
                f.write(self.cookie_var.get().strip())
        except:
            pass
            
        # 保存排除商品ID
        try:
            with open(self.exclude_ids_file, "w", encoding="utf-8") as f:
                f.write(self.exclude_product_ids_var.get().strip())
        except:
            pass
        
    def create_widgets(self):
        # 顶部框架：文件选择和配置
        top_frame = tk.Frame(self.root)
        top_frame.pack(fill="x", padx=10, pady=5)
        
        # 文件选择区域
        file_frame = tk.LabelFrame(top_frame, text="第一步：选择Excel文件", padx=10, pady=10)
        file_frame.pack(fill="x", pady=5)
        
        tk.Entry(file_frame, textvariable=self.file_path, width=60).pack(side="left", padx=5)
        tk.Button(file_frame, text="浏览...", command=self.browse_file).pack(side="left", padx=5)
        
        # Cookie 设置区域
        cookie_frame = tk.LabelFrame(top_frame, text="第二步：设置Cookie", padx=10, pady=10)
        cookie_frame.pack(fill="x", pady=5)
        
        tk.Label(cookie_frame, text="Cookie:").pack(side="left", padx=5)
        tk.Entry(cookie_frame, textvariable=self.cookie_var, width=80).pack(side="left", padx=5)
        
        # 配置区域
        config_frame = tk.LabelFrame(top_frame, text="第三步：配置", padx=10, pady=10)
        config_frame.pack(fill="x", pady=5)
        
        tk.Label(config_frame, text="线程数(并发查询数):").pack(side="left", padx=5)
        tk.Spinbox(config_frame, from_=1, to=20, textvariable=self.thread_count, width=5).pack(side="left", padx=5)
        tk.Label(config_frame, text="(建议4-8个，过多可能被封)").pack(side="left", padx=5)
        filter_frame = tk.Frame(config_frame)
        filter_frame.pack(fill="x", pady=6)
        tk.Label(filter_frame, text="排除商品id:").pack(side="left", padx=5)
        tk.Entry(filter_frame, textvariable=self.exclude_product_ids_var).pack(side="left", fill="x", expand=True, padx=5)
        
        # 操作区域
        action_frame = tk.Frame(self.root, padx=10, pady=5)
        action_frame.pack(fill="x", padx=10)
        
        self.start_btn = tk.Button(action_frame, text="开始处理", command=self.start_processing_thread, bg="#4CAF50", fg="white", font=("Arial", 12, "bold"), height=2)
        self.start_btn.pack(fill="x", padx=5)
        
        # 进度区域
        progress_frame = tk.LabelFrame(self.root, text="处理进度", padx=10, pady=5)
        progress_frame.pack(fill="x", padx=10, pady=5)
        
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.pack(fill="x", padx=5, pady=5)
        
        tk.Label(progress_frame, textvariable=self.status_var, fg="blue").pack(anchor="w", padx=5)
        
        # 日志区域
        log_frame = tk.LabelFrame(self.root, text="运行日志 (实时显示)", padx=10, pady=5)
        log_frame.pack(fill="both", expand=True, padx=10, pady=5)
        
        self.log_text = tk.Text(log_frame, height=15)
        self.log_text.pack(side="left", fill="both", expand=True)
        
        # 滚动条
        scrollbar = tk.Scrollbar(log_frame)
        scrollbar.pack(side="right", fill="y")
        self.log_text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.log_text.yview)

    def log(self, message):
        # 使用 after 方法确保在主线程更新 UI
        self.root.after(0, lambda: self._log(message))
        
    def _log(self, message):
        timestamp = time.strftime("%H:%M:%S", time.localtime())
        full_msg = f"[{timestamp}] {message}"
        self.log_text.insert("end", full_msg + "\n")
        self.log_text.see("end")

    def normalize_header(self, value):
        if value is None:
            return ""
        return str(value).strip().lower()

    def find_header_index(self, headers, candidates):
        normalized_map = {self.normalize_header(h): idx for idx, h in enumerate(headers)}
        for name in candidates:
            idx = normalized_map.get(self.normalize_header(name))
            if idx is not None:
                return idx
        return None

    def parse_excluded_product_ids(self):
        raw = self.exclude_product_ids_var.get().strip()
        if not raw:
            return set()
        parts = re.split(r"[,\s，]+", raw)
        return {p.strip() for p in parts if p.strip()}

    def to_float(self, value):
        if value is None:
            return 0.0
        text = str(value).strip()
        if not text:
            return 0.0
        text = text.replace(",", "").replace("，", "").replace("¥", "").replace("￥", "")
        try:
            return float(text)
        except:
            return 0.0

    def normalize_status(self, value):
        text = str(value).strip().replace(",", "，").replace(" ", "")
        return text

    def load_cost_map(self, data_file_path):
        base_dir = os.path.dirname(data_file_path)
        candidates = [
            os.path.join(base_dir, "成本表.xlsx"),
            os.path.join(os.getcwd(), "成本表.xlsx"),
            os.path.join(base_dir, "成本表.成本表.xlsx"),
            os.path.join(os.getcwd(), "成本表.成本表.xlsx")
        ]
        cost_file_path = ""
        for path in candidates:
            if os.path.exists(path):
                cost_file_path = path
                break
        if not cost_file_path:
            raise Exception("未找到成本表文件: 成本表.xlsx")
        self.log(f"加载成本表: {cost_file_path}")
        wb = openpyxl.load_workbook(cost_file_path, data_only=True)
        if "sku商品编码" in wb.sheetnames:
            ws = wb["sku商品编码"]
        else:
            ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [h if h is not None else "" for h in header_row]
        sku_col = self.find_header_index(headers, ["组合编码", "商家编码-规格维度"])
        cost_col = self.find_header_index(headers, ["sku成本"])
        if sku_col is None or cost_col is None:
            wb.close()
            raise Exception("成本表缺少必要表头: 组合编码 或 sku成本")
        cost_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            if sku_col >= len(row):
                continue
            sku_key = row[sku_col]
            if sku_key is None:
                continue
            sku_text = str(sku_key).strip()
            if not sku_text:
                continue
            cost_value = row[cost_col] if cost_col < len(row) else ""
            cost_map[sku_text] = self.to_float(cost_value)
        baby_powder_codes = set()
        if "母婴奶粉编码" in wb.sheetnames:
            baby_ws = wb["母婴奶粉编码"]
            for row in baby_ws.iter_rows(min_row=1, values_only=True):
                if not row:
                    continue
                max_cols = min(6, len(row))
                for idx in range(max_cols):
                    val = row[idx]
                    if val is None:
                        continue
                    text = str(val).strip()
                    if text:
                        baby_powder_codes.add(text)
        wb.close()
        self.log(f"成本表加载完成，共 {len(cost_map)} 条")
        return cost_map, baby_powder_codes

    def get_last_non_empty_col_csv(self, header):
        for idx in range(len(header) - 1, -1, -1):
            if str(header[idx]).strip():
                return idx
        return -1

    def get_last_non_empty_col_excel(self, ws):
        for col in range(ws.max_column, 0, -1):
            if str(ws.cell(row=1, column=col).value or "").strip():
                return col
        return 0

    def select_data_sheet(self, wb):
        candidates = []
        for ws in wb.worksheets:
            if ws.title in ("处理后的表", "汇总"):
                continue
            header_values = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
            non_empty_header_count = sum(1 for h in header_values if str(h or "").strip())
            data_rows = max(0, ws.max_row - 1)
            has_order_header = self.find_header_index(header_values, ["订单号"]) is not None
            score = (1 if has_order_header else 0, data_rows, non_empty_header_count)
            candidates.append((score, ws))
        if not candidates:
            return wb.active
        candidates.sort(key=lambda x: x[0], reverse=True)
        return candidates[0][1]

    def browse_file(self):
        filenames = filedialog.askopenfilenames(filetypes=[("Excel/CSV files", "*.xlsx *.csv"), ("Excel files", "*.xlsx"), ("CSV files", "*.csv")])
        if filenames:
            self.file_paths = list(filenames)
            if len(self.file_paths) <= 3:
                display_text = " | ".join(self.file_paths)
            else:
                display_text = " | ".join(self.file_paths[:3]) + f" ... 共{len(self.file_paths)}个文件"
            self.file_path.set(display_text)
            self.log(f"已选择 {len(self.file_paths)} 个文件")

    def start_processing_thread(self):
        if not self.file_paths:
            messagebox.showwarning("警告", "请先选择文件")
            return
            
        if not self.cookie_var.get().strip():
            messagebox.showwarning("警告", "请输入Cookie")
            return
            
        self.start_btn.config(state="disabled", text="正在处理中...")
        self.log_text.delete(1.0, "end") # 清空日志
        self.log("任务开始...")
        self.pending_files = self.file_paths.copy()
        self.total_files = len(self.pending_files)
        self.processed_files = 0
        self.last_output_file = None
        
        # 保存Cookie和排除ID配置
        self.save_config()
        
        # 开启线程
        threading.Thread(target=self.process_file, daemon=True).start()

    def process_row(self, row_data, cookie):
        row_idx, order_no = row_data
        
        if not order_no:
            return row_idx, "", "", False, "跳过"
            
        order_no_str = str(order_no).strip()
        result = get_charge_order_list(order_no_str, cookie)
        
        total_quoted_amount = 0.0
        fee_names = []
        msg = ""
        
        if result and "error" not in result:
            if result.get("success") and result.get("data"):
                data_list = result.get("data")
                if isinstance(data_list, list):
                    for item in data_list:
                        amount = item.get("quotedAmount")
                        if amount:
                            total_quoted_amount += float(amount)
                        fee_name = item.get("feeName")
                        if fee_name is not None and str(fee_name).strip():
                            fee_names.append(str(fee_name).strip())
                
                # 保留两位小数
                total_quoted_amount = round(total_quoted_amount, 2)
                msg = f"成功: {total_quoted_amount}"
            else:
                msg = "查询无数据"
        else:
            error_msg = result.get("error") if result else "未知错误"
            msg = f"失败: {error_msg}"
        fee_name_set = {name.strip() for name in fee_names if str(name).strip()}
        required_fees = {"基础服务费", "进口关税"}
        fee_ok = required_fees.issubset(fee_name_set)
        return row_idx, total_quoted_amount, ",".join(fee_names), fee_ok, msg

    def process_file(self):
        if not self.pending_files:
            self.root.after(0, lambda: self.finish_processing(self.last_output_file))
            return
        filename = self.pending_files.pop(0)
        self.processed_files += 1
        self.status_var.set("正在读取文件...")
        self.log(f"开始处理文件({self.processed_files}/{self.total_files}): {filename}")
        
        try:
            is_csv = filename.lower().endswith('.csv')
            rows = []
            wb = None
            ws = None
            source_ws = None
            csv_header = []
            excel_header = []
            excel_row_numbers = []
            
            is_csv_format = is_csv
            if is_csv:
                encodings = ['utf-8-sig', 'gbk', 'gb18030']
                encoding = None
                
                for enc in encodings:
                    try:
                        with open(filename, 'r', encoding=enc) as f:
                            f.read()
                        encoding = enc
                        break
                    except UnicodeDecodeError:
                        continue
                
                if not encoding:
                    raise Exception("无法识别文件编码，请尝试另存为UTF-8或GBK格式")
                    
                self.log(f"使用编码: {encoding}")
                
                with open(filename, 'r', encoding=encoding, newline='') as f:
                    reader = csv.reader(f)
                    all_rows = list(reader)
                    
                if not all_rows:
                    self.log("CSV文件为空")
                    self.last_output_file = filename
                    if self.pending_files:
                        self.process_file()
                        return
                    self.status_var.set("处理完成！")
                    self.root.after(0, lambda: self.finish_processing(self.last_output_file))
                    return
                    
                wb = openpyxl.Workbook()
                source_ws = wb.active
                source_ws.title = "Sheet1"
                for r_idx, row_data in enumerate(all_rows, 1):
                    for c_idx, val in enumerate(row_data, 1):
                        source_ws.cell(row=r_idx, column=c_idx, value=val)
                self.log("CSV文件已在内存中转换为Excel格式")
            else:
                wb = openpyxl.load_workbook(filename)
                source_ws = self.select_data_sheet(wb)
                self.log(f"使用工作表: {source_ws.title}")
                
            if "处理后的表" in wb.sheetnames:
                del wb["处理后的表"]
            ws = wb.copy_worksheet(source_ws)
            ws.title = "处理后的表"
            excel_header = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
            for row_idx in range(2, ws.max_row + 1):
                row_cells = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, max_col=ws.max_column))[0]
                rows.append(row_cells)
                excel_row_numbers.append(row_idx)
                
            is_csv = False  # 后续处理统一当做Excel
            
            total_rows = len(rows)
            
            if total_rows == 0:
                self.log("没有数据需要处理")
                self.last_output_file = filename
                if self.pending_files:
                    self.process_file()
                    return
                self.status_var.set("处理完成！")
                self.root.after(0, lambda: self.finish_processing(self.last_output_file))
                return

            self.log(f"共发现 {total_rows} 行数据")

            headers = excel_header
            order_no_idx = self.find_header_index(headers, ["订单号"])
            status_idx = self.find_header_index(headers, ["订单状态"])
            product_id_idx = self.find_header_index(headers, ["商品id", "商品ID"])
            sku_idx = self.find_header_index(headers, ["商家编码-规格维度"])
            qty_idx = self.find_header_index(headers, ["商品数量(件)", "商品数量"])
            receive_idx = self.find_header_index(headers, ["商家实收金额(元)", "商家实收金额"])

            if order_no_idx is None:
                raise Exception("未找到表头: 订单号")
            if status_idx is None:
                raise Exception("未找到表头: 订单状态")
            if sku_idx is None:
                raise Exception("未找到表头: 商家编码-规格维度")
            if qty_idx is None:
                raise Exception("未找到表头: 商品数量(件)")
            if receive_idx is None:
                raise Exception("未找到表头: 商家实收金额(元)")

            excluded_ids = self.parse_excluded_product_ids()
            if excluded_ids:
                self.log(f"排除商品id数量: {len(excluded_ids)}")

            cost_map, baby_powder_codes = self.load_cost_map(filename)

            results = {}
            tasks = []
            row_cache = {}
            special_row_flags = {}
            keep_row_flags = {}
            allowed_status = {"已发货，待收货", "已收货"}

            def get_row_value(row, idx):
                if idx is None:
                    return ""
                if idx >= len(row):
                    return ""
                if is_csv:
                    return row[idx]
                return row[idx].value

            for i, row in enumerate(rows):
                order_no = str(get_row_value(row, order_no_idx)).strip()
                order_status = self.normalize_status(get_row_value(row, status_idx))
                product_id = str(get_row_value(row, product_id_idx)).strip() if product_id_idx is not None else ""
                sku_text = str(get_row_value(row, sku_idx)).strip()
                qty_value = self.to_float(get_row_value(row, qty_idx))
                receive_value = self.to_float(get_row_value(row, receive_idx))

                if order_status not in allowed_status:
                    keep_row_flags[i] = False
                    continue
                if product_id and product_id in excluded_ids:
                    keep_row_flags[i] = False
                    continue
                keep_row_flags[i] = True
                if not order_no:
                    results[i] = {"query_result": "", "product_cost": "", "order_cost": "", "all_cost": "", "profit": "", "fee_names": "", "fee_ok": True, "cost_ok": True, "tax_ok": True}
                    continue

                row_cache[i] = {
                    "sku": sku_text,
                    "qty": qty_value,
                    "receive": receive_value
                }
                tasks.append((i, order_no))
                
            total_tasks = len(tasks)
            kept_rows = sum(1 for v in keep_row_flags.values() if v)
            removed_rows = total_rows - kept_rows
            self.log(f"状态和商品id筛选后保留 {kept_rows} 行，删除 {removed_rows} 行")
            self.log(f"共发现 {total_rows} 行数据，其中 {total_tasks} 行有效单号需要处理")
            
            processed_count = 0
            thread_num = self.thread_count.get()
            current_cookie = self.cookie_var.get().strip()
            
            if total_tasks > 0:
                self.log(f"启动 {thread_num} 个线程进行查询...")
                
                with ThreadPoolExecutor(max_workers=thread_num) as executor:
                    futures = {executor.submit(self.process_row, task, current_cookie): task for task in tasks}
                    
                    for future in as_completed(futures):
                        try:
                            row_idx, amount, fee_names, fee_ok, msg = future.result()
                            source_data = row_cache.get(row_idx, {})
                            sku_key = source_data.get("sku", "")
                            cost_ok = sku_key in cost_map
                            product_cost = round(cost_map.get(sku_key, 0.0), 2)
                            order_cost = round(product_cost * source_data.get("qty", 0.0), 2)
                            raw_query_result = round(self.to_float(amount), 2)
                            query_result = raw_query_result
                            is_special_order = False
                            tax_ok = True
                            if raw_query_result == 0:
                                if sku_key in baby_powder_codes:
                                    query_result = round(source_data.get("receive", 0.0) * 0.091, 2)
                                    is_special_order = True
                                else:
                                    tax_ok = False
                            all_cost = round(query_result + order_cost, 2)
                            profit = round(source_data.get("receive", 0.0) - all_cost, 2)
                            special_row_flags[row_idx] = is_special_order
                            results[row_idx] = {
                                "query_result": query_result,
                                "product_cost": product_cost,
                                "order_cost": order_cost,
                                "all_cost": all_cost,
                                "profit": profit,
                                "fee_names": fee_names,
                                "fee_ok": fee_ok,
                                "cost_ok": cost_ok,
                                "tax_ok": tax_ok
                            }
                        except Exception as e:
                            self.log(f"任务执行异常: {str(e)}")
                            continue
                        
                        processed_count += 1
                        progress = (processed_count / total_tasks) * 100
                        self.progress_var.set(progress)
                        self.status_var.set(f"进度: {processed_count}/{total_tasks}")
                        
                        # 打印日志 (每10条打印一次，避免刷屏)
                        if processed_count % 10 == 0 or processed_count == total_tasks:
                            self.log(f"已处理 {processed_count}/{total_tasks}")
            else:
                self.log("没有有效单号需要处理")
                self.progress_var.set(100)

            self.status_var.set("正在写入结果...")
            self.log("正在保存结果...")
            
            if is_csv_format:
                output_filename = filename[:-4] + "_处理结果.xlsx"
            else:
                output_filename = filename
                
            output_headers = ["查询结果", "商品成本", "订单成本", "综合成本", "毛利润", "费用项"]
            summary_totals = {
                "商品数量(件)": 0.0,
                "商家实收金额(元)": 0.0,
                "查询结果": 0.0,
                "订单成本": 0.0,
                "综合成本": 0.0,
                "毛利润": 0.0
            }
            
            existing_start_idx = self.find_header_index(excel_header, [output_headers[0]])
            if existing_start_idx is not None:
                start_col = existing_start_idx + 1
            else:
                start_col = self.get_last_non_empty_col_excel(ws) + 1
            for offset, header_name in enumerate(output_headers):
                ws.cell(row=1, column=start_col + offset, value=header_name)
            pink_fill = PatternFill(start_color="FFF6F8", end_color="FFF6F8", fill_type="solid")
            red_fill = PatternFill(start_color="FFFFE5E5", end_color="FFFFE5E5", fill_type="solid")
            white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
            gray_side = Side(style="thin", color="FFE7E6E6")
            gray_border = Border(left=gray_side, right=gray_side, top=gray_side, bottom=gray_side)
            empty_border = Border()
            error_rows_data = []
            for i, row in enumerate(rows):
                if not keep_row_flags.get(i, True):
                    continue
                row_result = results.get(i, {"query_result": "", "product_cost": "", "order_cost": "", "all_cost": "", "profit": "", "fee_names": "", "fee_ok": True, "cost_ok": True, "tax_ok": True})
                summary_totals["商品数量(件)"] += self.to_float(get_row_value(row, qty_idx))
                summary_totals["商家实收金额(元)"] += self.to_float(get_row_value(row, receive_idx))
                summary_totals["查询结果"] += self.to_float(row_result.get("query_result", 0))
                summary_totals["订单成本"] += self.to_float(row_result.get("order_cost", 0))
                summary_totals["综合成本"] += self.to_float(row_result.get("all_cost", 0))
                summary_totals["毛利润"] += self.to_float(row_result.get("profit", 0))
                current_row_idx = excel_row_numbers[i]
                ws.cell(row=current_row_idx, column=start_col + 0, value=row_result.get("query_result", ""))
                ws.cell(row=current_row_idx, column=start_col + 1, value=row_result.get("product_cost", ""))
                ws.cell(row=current_row_idx, column=start_col + 2, value=row_result.get("order_cost", ""))
                ws.cell(row=current_row_idx, column=start_col + 3, value=row_result.get("all_cost", ""))
                ws.cell(row=current_row_idx, column=start_col + 4, value=row_result.get("profit", ""))
                ws.cell(row=current_row_idx, column=start_col + 5, value=row_result.get("fee_names", ""))
                fee_issue = (not row_result.get("fee_ok", True)) and (not special_row_flags.get(i, False))
                if fee_issue or (not row_result.get("cost_ok", True)) or (not row_result.get("tax_ok", True)):
                    row_fill = red_fill
                elif special_row_flags.get(i, False):
                    row_fill = pink_fill
                else:
                    row_fill = white_fill
                for col_no in range(1, ws.max_column + 1):
                    cell = ws.cell(row=current_row_idx, column=col_no)
                    cell.fill = row_fill
                    if row_fill == white_fill:
                        cell.border = gray_border
                    else:
                        cell.border = empty_border
                if row_fill == red_fill:
                    error_row_values = [ws.cell(row=current_row_idx, column=col_no).value for col_no in range(1, ws.max_column + 1)]
                    error_rows_data.append(error_row_values)
            delete_row_numbers = [excel_row_numbers[i] for i in range(len(rows)) if not keep_row_flags.get(i, True)]
            for row_no in sorted(delete_row_numbers, reverse=True):
                ws.delete_rows(row_no, 1)
            if "汇总" in wb.sheetnames:
                del wb["汇总"]
            if "错误项" in wb.sheetnames:
                del wb["错误项"]
            error_ws = None
            if error_rows_data:
                error_ws = wb.create_sheet("错误项")
                error_header_values = [ws.cell(row=1, column=col_no).value for col_no in range(1, ws.max_column + 1)]
                error_ws.append(error_header_values)
                for row_values in error_rows_data:
                    error_ws.append(row_values)
            summary_ws = wb.create_sheet("汇总")
            summary_fields = ["商家实收金额(元)", "商品数量(件)", "查询结果", "订单成本", "综合成本", "毛利润"]
            for idx, field_name in enumerate(summary_fields, start=1):
                summary_ws.cell(row=1, column=idx, value=field_name)
                summary_ws.cell(row=2, column=idx, value=round(summary_totals[field_name], 2))
            if source_ws is not None:
                ordered_sheets = [source_ws, ws]
                if error_ws is not None:
                    ordered_sheets.append(error_ws)
                ordered_sheets.append(summary_ws)
                remaining_sheets = [s for s in wb.worksheets if s not in ordered_sheets]
                wb._sheets = ordered_sheets + remaining_sheets
            
            wb.save(output_filename)
                
            self.log(f"文件已保存: {output_filename}")
            self.last_output_file = output_filename
            if self.pending_files:
                self.process_file()
                return
            self.status_var.set("处理完成！")
            self.root.after(0, lambda: self.finish_processing(self.last_output_file))
            
        except Exception as e:
            self.log(f"发生错误: {str(e)}")
            import traceback
            traceback.print_exc()
            if self.pending_files:
                self.process_file()
                return
            self.root.after(0, lambda: self.finish_processing(self.last_output_file))

    def finish_processing(self, output_file):
        self.start_btn.config(state="normal", text="开始处理")
        if output_file:
            self.log(f"任务完成！结果已更新至原文件: {output_file}")
            # 不弹出窗口
        else:
            self.log("任务失败，请检查日志。")
            # 不弹出窗口

if __name__ == "__main__":
    root = tk.Tk()
    app = BatchQueryApp(root)
    root.mainloop()
