# -*- coding: utf-8 -*-
"""
利润核算引擎 (ProfitEngine)
从 gui_batch_query.py 抽取的纯核算逻辑，无任何 UI 依赖。
- 日志/进度通过回调注入
- 配置表/订单文件路径全部参数化
- 与原 GUI 逻辑保持一致，差异点用 [Web版修复] 标注

修复清单（对照 docs/利润核算Web服务-产品规划.md 风险登记册）:
- R2: 销售计划登记表 store_link_map 残留变量导致只统计最后一个店铺
- R3a: 空输入(无有效订单)时报错终止，而非静默生成空文件
- R3b: 文件名日期识别失败时告警（不再静默用"今天"兜底）
- R3c: 多天合并时校验同店铺表头一致性，不一致时告警
"""
import csv
import difflib
import os
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import openpyxl
import requests
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter

# 接口地址
API_URL = "https://b.cainiao.com/merchantchargeorder/queryChargeOrderListNew"
KNOWN_STORES = ["庆余", "雅丽丹", "大咖猴", "趣味猴", "品味"]
REQUIRED_FEES = {"基础服务费", "进口关税"}


class EngineError(Exception):
    """引擎业务错误，带错误码（2xxx 配置 / 4xxx 引擎内部）"""

    def __init__(self, message, code=40000, details=None):
        super().__init__(message)
        self.code = code
        self.details = details or {}


def get_charge_order_list(order_no, cookie):
    headers = {
        "accept": "application/json, text/plain, */*",
        "bx-v": "2.5.36",
        "content-type": "application/json",
        "h-csrf": "635e77ca-003a-498a-9a58-18c213306b08",
        "referer": "https://b.cainiao.com/cf_seller/charge-order/charge-order-query",
        "sec-ch-ua": '"Not:A-Brand";v="99", "Microsoft Edge";v="145", "Chromium";v="145"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36 Edg/145.0.0.0",
        "Cookie": cookie
    }
    post_data = {
        "orderNo": order_no,
        "currentPage": 1,
        "pageSize": 10,
        "orderNoType": 0
    }
    try:
        response = requests.post(
            url=API_URL,
            headers=headers,
            data=json_dumps(post_data),
            timeout=30,
            allow_redirects=False
        )
        if response.status_code in (301, 302, 303, 307, 308):
            return {"error": "请求被重定向 (Cookie可能失效)"}
        if response.status_code != 200:
            return {"error": f"请求失败: 状态码 {response.status_code}"}
        return response.json()
    except Exception as e:
        return {"error": f"请求异常: {e}"}


def json_dumps(obj):
    import json
    return json.dumps(obj)


class ProfitEngine:
    def __init__(self, log_cb=None, progress_cb=None):
        """
        :param log_cb: 日志回调 log_cb(msg)
        :param progress_cb: 进度回调 progress_cb(pct, text)，pct 为 0-100 浮点
        """
        self._log_cb = log_cb or (lambda msg: None)
        self._progress_cb = progress_cb or (lambda pct, text: None)
        self.tax_cache = {}
        self.tax_cache_dirty = False

    # ---------- 回调 ----------
    def log(self, msg):
        self._log_cb(msg)

    def progress(self, pct, text):
        self._progress_cb(pct, text)

    # ---------- 工具方法 ----------
    def normalize_header(self, value):
        if value is None:
            return ""
        return str(value).strip().lower()

    def find_header_index(self, headers, candidates):
        normalized_map = {self.normalize_header(h): idx for idx, h in enumerate(headers)}
        for name in candidates:
            if self.normalize_header(name) in normalized_map:
                return normalized_map[self.normalize_header(name)]
        return None

    def to_float(self, value):
        if value is None:
            return 0.0
        text = str(value).strip().replace(",", "").replace("，", "").replace("¥", "").replace("￥", "")
        try:
            return float(text)
        except Exception:
            return 0.0

    def clean_id(self, val):
        if val is None:
            return ""
        v_str = str(val).strip()
        if v_str.endswith(".0"):
            return v_str[:-2]
        return v_str

    def parse_excluded_product_ids(self, raw):
        if not raw:
            return set()
        return {p.strip() for p in re.split(r"[,\s，]+", raw) if p.strip()}

    def parse_cost_date_range(self, date_str):
        if not date_str:
            return None, None
        date_str = str(date_str).strip()
        parts = date_str.split('-')
        if len(parts) != 2:
            return None, None

        start_str, end_str = parts[0].strip(), parts[1].strip()

        def parse_single_date(s):
            if not s or s == "至今":
                return None
            s = s.replace('.', '/').replace('-', '/')
            try:
                p = s.split('/')
                if len(p) == 3:
                    return datetime(int(p[0]), int(p[1]), int(p[2])).date()
            except Exception:
                pass
            return None

        return parse_single_date(start_str), parse_single_date(end_str)

    def parse_date(self, date_val):
        if not date_val:
            return None
        if isinstance(date_val, datetime):
            return date_val.date()
        s = str(date_val).strip()
        match = re.search(r'(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})', s)
        if match:
            return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3))).date()
        return None

    def get_sort_key(self, p_name, order_list):
        if not order_list:
            return (1, p_name)

        t_clean = p_name.lower().replace(" ", "")
        best_idx = len(order_list)
        best_ratio = 0
        for i, o in enumerate(order_list):
            o_clean = o.lower().replace(" ", "")
            if t_clean in o_clean or o_clean in t_clean:
                return (0, i)
            ratio = difflib.SequenceMatcher(None, t_clean, o_clean).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_idx = i

        if best_ratio > 0.5:
            return (0, best_idx)
        return (1, p_name)

    def extract_store_name(self, filename):
        base = os.path.basename(filename)
        for s in KNOWN_STORES:
            if s in base:
                return s
        return "未知店铺"

    def extract_date_from_filename(self, filename):
        base = os.path.basename(filename)

        # 完整格式 20260526 / 2026-05-26 / 2026年05月26日
        match = re.search(r'(20\d{2})[-年]?(\d{1,2})[-月]?(\d{1,2})日?', base)
        if match:
            year, month, day = match.groups()
            return f"{year}年{int(month):02d}月{int(day):02d}日"

        # 简写格式 26-5-26 (默认加 20 前缀)
        match = re.search(r'(\d{2})[-年](\d{1,2})[-月](\d{1,2})日?', base)
        if match:
            year_short, month, day = match.groups()
            return f"20{year_short}年{int(month):02d}月{int(day):02d}日"

        # 纯数字简写 260526
        match = re.search(r'(2\d)(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])', base)
        if match:
            year_short, month, day = match.groups()
            return f"20{year_short}年{month}月{day}日"

        return ""

    def read_source_file(self, filename):
        is_csv = filename.lower().endswith('.csv')
        rows = []
        headers = []
        if is_csv:
            encodings = ['utf-8-sig', 'gbk', 'gb18030']
            encoding = None
            for enc in encodings:
                try:
                    with open(filename, 'r', encoding=enc) as f:
                        f.read()
                    encoding = enc
                    break
                except UnicodeDecodeError:
                    continue
            if not encoding:
                raise EngineError(f"无法识别文件编码: {filename}", code=21001)

            with open(filename, 'r', encoding=encoding, newline='') as f:
                all_rows = list(csv.reader(f))
            if all_rows:
                headers = all_rows[0]
                rows = all_rows[1:]
        else:
            try:
                wb = openpyxl.load_workbook(filename, data_only=True)
            except Exception as e:
                raise EngineError(f"无法打开 Excel 文件: {filename} ({e})", code=21002)
            ws = wb.active
            for ws_test in wb.worksheets:
                if ws_test.title not in ("处理后的表", "汇总", "错误项"):
                    ws = ws_test
                    break

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c or "") for c in row]
                else:
                    rows.append([c for c in row])
            wb.close()
        return headers, rows

    # ---------- 配置加载 ----------
    def load_cost_map(self, cost_path):
        if not cost_path or not os.path.exists(cost_path):
            raise EngineError("未找到成本表文件: 成本表.xlsx，请上传后重试", code=20101)
        self.log(f"加载成本表: {cost_path}")
        try:
            wb = openpyxl.load_workbook(cost_path, data_only=True)
        except Exception as e:
            raise EngineError(f"成本表无法打开: {cost_path} ({e})", code=20102)
        ws = wb["sku商品编码"] if "sku商品编码" in wb.sheetnames else wb.active

        headers = [str(h or "") for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        sku_col = self.find_header_index(headers, ["组合编码", "商家编码-规格维度"])
        cost_col = self.find_header_index(headers, ["sku成本"])
        date_col = self.find_header_index(headers, ["日期", "时间"])

        if sku_col is None or cost_col is None:
            wb.close()
            raise EngineError(f"成本表缺少必要表头: 组合编码 或 sku成本 (当前表头: {[h for h in headers if h]})", code=20103)

        cost_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or sku_col >= len(row) or row[sku_col] is None:
                continue
            sku_text = str(row[sku_col]).strip()
            if not sku_text:
                continue
            cost_value = row[cost_col] if cost_col < len(row) else ""

            date_val = ""
            if date_col is not None and date_col < len(row):
                date_val = row[date_col]
            elif len(row) > 6 and row[6]:
                date_val = row[6]

            s_date, e_date = None, None
            if date_val:
                s_date, e_date = self.parse_cost_date_range(date_val)

            if sku_text not in cost_map:
                cost_map[sku_text] = []

            cost_map[sku_text].append({
                "start": s_date,
                "end": e_date,
                "cost": self.to_float(cost_value)
            })

        baby_powder_codes = set()
        if "母婴奶粉编码" in wb.sheetnames:
            baby_ws = wb["母婴奶粉编码"]
            for row in baby_ws.iter_rows(min_row=1, values_only=True):
                if not row:
                    continue
                for idx in range(min(6, len(row))):
                    if row[idx] is not None and str(row[idx]).strip():
                        baby_powder_codes.add(str(row[idx]).strip())
        wb.close()
        self.log(f"成本表加载完成，共 {len(cost_map)} 条")
        return cost_map, baby_powder_codes

    def load_link_map(self, link_path):
        if not link_path or not os.path.exists(link_path):
            raise EngineError("无效的产品链接汇总表路径，请上传后重试", code=20111)
        self.log(f"加载产品链接汇总表: {link_path}")
        try:
            wb = openpyxl.load_workbook(link_path, data_only=True)
        except Exception as e:
            raise EngineError(f"产品链接汇总表无法打开: {link_path} ({e})", code=20112)
        ws = wb.active
        headers = [str(h or "") for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

        store_col = self.find_header_index(headers, ["店铺名称", "店铺"])
        prod_col = self.find_header_index(headers, ["产品名称", "产品"])
        id_col = self.find_header_index(headers, ["商品id", "商品ID", "链接id"])

        if store_col is None or prod_col is None or id_col is None:
            wb.close()
            raise EngineError(f"产品链接汇总表缺少表头: 店铺名称 | 产品名称 | 商品ID (当前表头: {[h for h in headers if h]})", code=20113)

        link_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or store_col >= len(row) or row[store_col] is None:
                continue
            store_name = str(row[store_col]).strip()
            prod_name = str(row[prod_col]).strip() if prod_col < len(row) and row[prod_col] is not None else "未知产品"
            prod_id = self.clean_id(row[id_col]) if id_col < len(row) else ""

            if not prod_id:
                continue

            if store_name not in link_map:
                link_map[store_name] = {}
            link_map[store_name][prod_id] = prod_name

        wb.close()
        self.log(f"产品链接汇总表加载完成，包含 {len(link_map)} 个店铺的映射")
        return link_map

    def load_promo_map(self, promo_paths):
        if not promo_paths:
            raise EngineError("请选择至少一个推广报表文件", code=20121)

        promo_map = {}
        total_loaded = 0

        for promo_path in promo_paths:
            if not promo_path or not os.path.exists(promo_path):
                continue

            store_name = self.extract_store_name(promo_path)
            self.log(f"加载推广报表: {os.path.basename(promo_path)} (识别店铺: {store_name})")
            try:
                wb = openpyxl.load_workbook(promo_path, data_only=True)
            except Exception as e:
                raise EngineError(f"推广报表无法打开: {promo_path} ({e})", code=20122)
            ws = wb.active
            headers = [str(h or "") for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

            id_col = self.find_header_index(headers, ["商品ID", "商品id", "链接id"])
            cost_col = self.find_header_index(headers, ["总花费(元)", "花费", "总花费"])
            date_col = self.find_header_index(headers, ["日期", "时间"])

            if id_col is None or cost_col is None:
                wb.close()
                self.log(f"警告: 推广报表 {os.path.basename(promo_path)} 缺少必要表头(商品ID/总花费)，已跳过")
                continue

            if store_name not in promo_map:
                promo_map[store_name] = {}

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or id_col >= len(row) or row[id_col] is None:
                    continue
                prod_id = self.clean_id(row[id_col])
                cost_val = self.to_float(row[cost_col]) if cost_col < len(row) else 0.0

                row_date_str = "未指定日期"
                if date_col is not None and date_col < len(row) and row[date_col]:
                    date_val = row[date_col]
                    date_obj = self.parse_date(date_val)
                    if date_obj:
                        row_date_str = f"{date_obj.year}年{date_obj.month:02d}月{date_obj.day:02d}日"
                    else:
                        row_date_str = str(date_val).strip()
                else:
                    file_date = self.extract_date_from_filename(promo_path)
                    if file_date:
                        row_date_str = file_date

                if prod_id:
                    if prod_id not in promo_map[store_name]:
                        promo_map[store_name][prod_id] = {"total": 0.0, "daily": {}}

                    promo_map[store_name][prod_id]["total"] += cost_val
                    if row_date_str != "未指定日期":
                        promo_map[store_name][prod_id]["daily"][row_date_str] = promo_map[store_name][prod_id]["daily"].get(row_date_str, 0.0) + cost_val
                    total_loaded += 1

            wb.close()

        self.log(f"推广报表全部加载完成，共解析了 {total_loaded} 条推广数据")
        return promo_map

    def load_subsidy_map(self, subsidy_path):
        if not subsidy_path or not os.path.exists(subsidy_path):
            self.log("未选择官补映射表，将跳过官补计算")
            return {}

        self.log(f"加载官补映射表: {subsidy_path}")
        try:
            wb = openpyxl.load_workbook(subsidy_path, data_only=True)
        except Exception as e:
            raise EngineError(f"官补映射表无法打开: {subsidy_path} ({e})", code=20132)
        ws = wb.active
        headers = [str(h or "") for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

        link_id_col = self.find_header_index(headers, ["链接id", "商品ID"])
        sku_col = self.find_header_index(headers, ["商家编码-规格维度", "规格编码"])
        subsidy_col = self.find_header_index(headers, ["官补金额"])
        start_col = self.find_header_index(headers, ["起始日期", "开始日期"])
        end_col = self.find_header_index(headers, ["结束日期"])

        if link_id_col is None or sku_col is None or subsidy_col is None:
            wb.close()
            self.log("警告: 官补映射表缺少必要表头(链接id/商家编码-规格维度/官补金额)，已跳过")
            return {}

        subsidy_map = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or link_id_col >= len(row) or row[link_id_col] is None:
                continue

            link_id = self.clean_id(row[link_id_col])
            sku = str(row[sku_col]).strip() if sku_col < len(row) and row[sku_col] is not None else ""
            subsidy_val = self.to_float(row[subsidy_col]) if subsidy_col < len(row) else 0.0

            start_date = None
            end_date = None
            if start_col is not None and start_col < len(row) and row[start_col]:
                start_date = self.parse_date(row[start_col])
            if end_col is not None and end_col < len(row) and row[end_col]:
                end_date = self.parse_date(row[end_col])

            if link_id and sku:
                if link_id not in subsidy_map:
                    subsidy_map[link_id] = {}
                if sku not in subsidy_map[link_id]:
                    subsidy_map[link_id][sku] = []
                subsidy_map[link_id][sku].append({
                    "start": start_date,
                    "end": end_date,
                    "amount": subsidy_val
                })

        wb.close()
        self.log(f"官补映射表加载完成，包含 {len(subsidy_map)} 个链接的官补规则")
        return subsidy_map

    def load_sort_list(self, sort_path):
        if not sort_path or not os.path.exists(sort_path):
            self.log("未选择或找不到产品排序表，将使用默认排序")
            return []

        self.log(f"加载产品排序表: {sort_path}")
        try:
            wb = openpyxl.load_workbook(sort_path, data_only=True)
        except Exception as e:
            raise EngineError(f"产品排序表无法打开: {sort_path} ({e})", code=20142)
        ws = wb.active
        order_list = []
        for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
            if row and row[0]:
                val = str(row[0]).strip()
                if val:
                    order_list.append(val)
        wb.close()
        self.log(f"成功加载 {len(order_list)} 个排序规则")
        return order_list

    # ---------- 税运查询 ----------
    def process_row_api(self, task, cookie, enable_tax=True):
        row_id, order_no = task
        if not order_no:
            return row_id, 0.0, {}, False, "跳过"

        if not enable_tax:
            return row_id, 0.0, {}, True, "不查税运"

        if str(order_no).strip() in self.tax_cache:
            cached = self.tax_cache[str(order_no).strip()]
            return row_id, cached.get("amount", 0.0), cached.get("fee_details", {}), cached.get("fee_ok", True), "从缓存读取"

        result = None
        max_retries = 3
        for attempt in range(max_retries):
            result = get_charge_order_list(str(order_no).strip(), cookie)
            if result and "error" not in result and result.get("success") and result.get("data"):
                break
            time.sleep(0.5)

        total_quoted = 0.0
        fee_details = {}
        msg = ""

        if result and "error" not in result:
            if result.get("success") and result.get("data"):
                data_list = result.get("data")
                if isinstance(data_list, list):
                    for item in data_list:
                        amount = item.get("quotedAmount")
                        fee_name = item.get("feeName")

                        val = 0.0
                        if amount:
                            val = float(amount)
                        total_quoted += val

                        if fee_name and str(fee_name).strip():
                            fn = str(fee_name).strip()
                            fee_details[fn] = fee_details.get(fn, 0.0) + val

                total_quoted = round(total_quoted, 2)
                msg = f"成功: {total_quoted}"
            else:
                msg = "查询无数据"
        else:
            msg = f"失败: {result.get('error') if result else '未知错误'}"

        fee_ok = REQUIRED_FEES.issubset(set(fee_details.keys()))
        return row_id, total_quoted, fee_details, fee_ok, msg

    # ---------- 主流程 ----------
    def process(self, order_files, configs, cookie, thread_count=2, enable_tax=True,
                exclude_ids="", output_dir=None):
        """
        :param order_files: 订单源文件路径列表 (CSV/XLSX)
        :param configs: {"cost": path, "link": path, "promo": [paths], "subsidy": path|None, "sort": path|None}
        :param cookie: 菜鸟 API Cookie
        :param thread_count: 税运查询并发线程数
        :param enable_tax: 是否查询税运
        :param exclude_ids: 排除商品ID (英文逗号分隔)
        :param output_dir: 输出目录；None 时取第一个订单文件所在目录
        :return: {"files": [绝对路径], "stats": {...}, "tax_cache": {...}}
        """
        if not order_files:
            raise EngineError("未选择任何业务源数据文件", code=21011)

        cost_map, baby_powder_codes = self.load_cost_map(configs.get("cost"))
        link_map = self.load_link_map(configs.get("link"))
        promo_map = self.load_promo_map(configs.get("promo") or [])
        subsidy_map = self.load_subsidy_map(configs.get("subsidy"))
        order_list = self.load_sort_list(configs.get("sort"))
        excluded_ids = self.parse_excluded_product_ids(exclude_ids)

        all_date_store_data = {}
        all_tasks = []

        allowed_status = {"已发货，待收货", "已收货"}
        task_counter = 0
        raw_total = 0

        # 阶段 1：读取所有文件并构建查询任务
        for filepath in order_files:
            store_name = self.extract_store_name(filepath)
            file_date = self.extract_date_from_filename(filepath)
            if not file_date:
                self.log(f"警告: 无法从文件名识别日期: {os.path.basename(filepath)}，将使用今天作为归属日期")
                file_date = datetime.now().strftime("%Y年%m月%d日")
            self.log(f"正在读取: {os.path.basename(filepath)} (识别店铺: {store_name})")
            headers, rows = self.read_source_file(filepath)

            order_no_idx = self.find_header_index(headers, ["订单号"])
            status_idx = self.find_header_index(headers, ["订单状态"])
            product_id_idx = self.find_header_index(headers, ["商品id", "商品ID"])
            sku_idx = self.find_header_index(headers, ["商家编码-规格维度"])
            qty_idx = self.find_header_index(headers, ["商品数量(件)", "商品数量"])
            receive_idx = self.find_header_index(headers, ["商家实收金额(元)", "商家实收金额"])
            discount_idx = self.find_header_index(headers, ["平台优惠折扣(元)", "平台优惠折扣", "平台优惠"])
            pay_time_idx = self.find_header_index(headers, ["支付时间", "付款时间", "订单付款时间"])

            if None in (order_no_idx, status_idx, sku_idx, qty_idx, receive_idx):
                self.log(f"跳过文件 {os.path.basename(filepath)}: 缺少必要表头 (订单号/订单状态/商家编码-规格维度/商品数量/商家实收金额)")
                continue

            for i, row in enumerate(rows):
                raw_total += 1
                order_status = str(row[status_idx] if status_idx < len(row) else "").strip().replace(",", "，").replace(" ", "")
                product_id = self.clean_id(row[product_id_idx] if product_id_idx is not None and product_id_idx < len(row) else None)

                row_date = file_date
                row_date_obj = None
                if pay_time_idx is not None and pay_time_idx < len(row) and row[pay_time_idx]:
                    pay_time_str = str(row[pay_time_idx]).strip()
                    match = re.search(r'(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})', pay_time_str)
                    if match:
                        y, m, d = match.groups()
                        row_date = f"{y}年{int(m):02d}月{int(d):02d}日"
                        row_date_obj = datetime(int(y), int(m), int(d)).date()

                if not row_date_obj:
                    match = re.search(r'(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})', row_date)
                    if match:
                        row_date_obj = datetime(int(match.group(1)), int(match.group(2)), int(match.group(3))).date()

                if row_date not in all_date_store_data:
                    all_date_store_data[row_date] = {}
                if store_name not in all_date_store_data[row_date]:
                    all_date_store_data[row_date][store_name] = {"headers": headers, "raw_rows": [], "rows": [], "results": {}}

                all_date_store_data[row_date][store_name]["raw_rows"].append(row)

                if order_status not in allowed_status:
                    continue
                if product_id and product_id in excluded_ids:
                    continue

                order_no = str(row[order_no_idx] if order_no_idx < len(row) else "").strip()
                sku_text = str(row[sku_idx] if sku_idx < len(row) else "").strip()
                qty_val = self.to_float(row[qty_idx] if qty_idx < len(row) else 0)
                receive_val = self.to_float(row[receive_idx] if receive_idx < len(row) else 0)
                discount_val = self.to_float(row[discount_idx] if discount_idx is not None and discount_idx < len(row) else 0)

                subsidy_val = 0.0
                if product_id in subsidy_map and sku_text in subsidy_map[product_id]:
                    rules = subsidy_map[product_id][sku_text]
                    for rule in rules:
                        s = rule["start"]
                        e = rule["end"]
                        if (not s or (row_date_obj and row_date_obj >= s)) and \
                           (not e or (row_date_obj and row_date_obj <= e)):
                            subsidy_val = rule["amount"]
                            break

                task_id = None
                if order_no:
                    task_id = f"{row_date}_{store_name}_{len(all_date_store_data[row_date][store_name]['rows'])}"
                    all_tasks.append((task_id, order_no))
                    all_date_store_data[row_date][store_name]["results"][task_id] = {
                        "sku": sku_text, "qty": qty_val, "receive": receive_val, "discount": discount_val,
                        "subsidy": subsidy_val, "product_id": product_id, "original_row": row, "date_obj": row_date_obj
                    }

                all_date_store_data[row_date][store_name]["rows"].append({
                    "task_id": task_id,
                    "original_row": row
                })

        total_tasks = len(all_tasks)
        self.log(f"总计需要查询 {total_tasks} 个有效订单 (共读取 {raw_total} 行原始数据)")

        # [Web版修复 R3a] 空输入不再静默生成空文件
        if total_tasks == 0:
            if not all_date_store_data:
                raise EngineError("未找到任何有效订单数据：所有文件缺少必要表头或无数据行。请检查订单源文件格式", code=21012)
            self.log("警告: 没有符合条件的订单(状态过滤/排除ID过滤后为空)，将仅生成原表数据")
            filtered_only = True
        else:
            filtered_only = False

        global_fee_names = set()
        processed_count = 0
        thread_num = max(1, min(int(thread_count), 4))  # Web版：并发上限 4

        # 阶段 2：并发查询 API
        if total_tasks > 0:
            with ThreadPoolExecutor(max_workers=thread_num) as executor:
                futures = {executor.submit(self.process_row_api, task, cookie, enable_tax): task for task in all_tasks}
                for future in as_completed(futures):
                    try:
                        task_id, amount, fee_details, fee_ok, msg = future.result()
                        parts = task_id.split("_", 2)
                        row_date = parts[0]
                        store_name = parts[1]
                        res_data = all_date_store_data[row_date][store_name]["results"][task_id]

                        for fn in fee_details.keys():
                            global_fee_names.add(fn)

                        sku_key = res_data["sku"]
                        order_date_obj = res_data.get("date_obj")

                        product_cost = 0.0
                        cost_ok = False

                        if sku_key in cost_map:
                            cost_ok = True
                            rules = cost_map[sku_key]
                            matched_cost = None
                            fallback_cost = None

                            for rule in rules:
                                s = rule["start"]
                                e = rule["end"]

                                if not s and not e:
                                    fallback_cost = rule["cost"]
                                    continue

                                if (not s or (order_date_obj and order_date_obj >= s)) and \
                                   (not e or (order_date_obj and order_date_obj <= e)):
                                    matched_cost = rule["cost"]
                                    break

                            if matched_cost is not None:
                                product_cost = matched_cost
                            elif fallback_cost is not None:
                                product_cost = fallback_cost
                            elif rules:
                                product_cost = rules[0]["cost"]

                        product_cost = round(product_cost, 2)
                        order_cost = round(product_cost * res_data["qty"], 2)

                        raw_query = round(self.to_float(amount), 2)
                        query_result = raw_query
                        is_special = False
                        tax_ok = True

                        if raw_query == 0:
                            if sku_key in baby_powder_codes:
                                query_result = round(res_data["receive"] * 0.091, 2)
                                is_special = True
                            else:
                                tax_ok = False

                        all_cost = round(query_result + order_cost, 2)

                        platform_fee = round(res_data["receive"] * 0.006, 2)

                        profit = round(res_data["receive"] - all_cost - platform_fee + res_data["subsidy"], 2)

                        res_data.update({
                            "query_result": query_result,
                            "product_cost": product_cost,
                            "order_cost": order_cost,
                            "all_cost": all_cost,
                            "platform_fee": platform_fee,
                            "profit": profit,
                            "fee_details": fee_details,
                            "fee_ok": fee_ok,
                            "cost_ok": cost_ok,
                            "tax_ok": tax_ok,
                            "is_special": is_special
                        })

                        task_order_no = str(futures[future][1]).strip()
                        fee_issue = not fee_ok and not is_special
                        is_red = fee_issue or not cost_ok or not tax_ok

                        if not is_red and task_order_no and task_order_no not in self.tax_cache:
                            self.tax_cache[task_order_no] = {
                                "amount": amount,
                                "fee_details": fee_details,
                                "fee_ok": fee_ok,
                                "msg": msg
                            }
                            self.tax_cache_dirty = True

                    except Exception as e:
                        self.log(f"查询异常: {str(e)}")

                    processed_count += 1
                    self.progress((processed_count / total_tasks) * 100, f"查询进度: {processed_count}/{total_tasks}")
                    if processed_count % 20 == 0 or processed_count == total_tasks:
                        self.log(f"已查询 {processed_count}/{total_tasks}")

        # 多天合并
        if len(all_date_store_data) > 1:
            sorted_dates = sorted(list(all_date_store_data.keys()))
            min_date = sorted_dates[0]
            max_date = sorted_dates[-1]
            combined_date_str = f"{min_date}至{max_date}"

            combined_store_data = {}
            for d_str in sorted_dates:
                stores_data = all_date_store_data[d_str]
                for store_name, s_data in stores_data.items():
                    if store_name not in combined_store_data:
                        combined_store_data[store_name] = {
                            "headers": s_data["headers"],
                            "raw_rows": [],
                            "rows": [],
                            "results": {}
                        }
                    else:
                        # [Web版修复 R3c] 多天同店铺表头不一致时告警
                        if s_data["headers"] != combined_store_data[store_name]["headers"]:
                            self.log(f"警告: 店铺 {store_name} 在 {d_str} 与之前日期的表头不一致，合并时按首日表头对齐，可能导致数据错位")
                    combined_store_data[store_name]["raw_rows"].extend(s_data.get("raw_rows", []))
                    combined_store_data[store_name]["rows"].extend(s_data["rows"])
                    combined_store_data[store_name]["results"].update(s_data["results"])

            all_date_store_data[combined_date_str] = combined_store_data

        # 阶段 3：生成最终 Excel
        self.log("正在生成汇总 Excel...")
        self.progress(100, "正在生成汇总文件...")

        if output_dir is None:
            output_dir = os.path.dirname(order_files[0])
        os.makedirs(output_dir, exist_ok=True)

        result_files = []
        pink_fill = PatternFill(start_color="FFF6F8", end_color="FFF6F8", fill_type="solid")
        red_fill = PatternFill(start_color="FFFFE5E5", end_color="FFFFE5E5", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
        gray_side = Side(style="thin", color="FFE7E6E6")
        gray_fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
        gray_border = Border(left=gray_side, right=gray_side, top=gray_side, bottom=gray_side)

        for date_str, all_store_data in all_date_store_data.items():
            sorted_fee_names = sorted(list(global_fee_names))
            included_dates = [d for d in all_date_store_data.keys() if "至" not in d]

            def get_promo(s_name, p_id):
                p_data = promo_map.get(s_name, {}).get(p_id, {})
                if "至" in date_str:
                    return sum(p_data.get("daily", {}).get(d, 0.0) for d in included_dates)
                else:
                    return p_data.get("daily", {}).get(date_str, 0.0)

            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            global_agg = {}
            global_store_totals = {}
            store_agg_map = {}

            sheet_groups = {
                "raw": [],
                "proc": [],
                "detail": []
            }

            for store_name, data in all_store_data.items():
                store_headers = data["headers"]

                def convert_row_types(row):
                    new_row = []
                    for val in row:
                        if isinstance(val, str):
                            v_strip = val.strip()
                            if re.match(r'^-?\d+\.\d+$', v_strip):
                                new_row.append(float(v_strip))
                            elif re.match(r'^-?\d+$', v_strip):
                                if len(v_strip) <= 15:
                                    new_row.append(int(v_strip))
                                else:
                                    new_row.append(v_strip)
                            else:
                                new_row.append(val)
                        else:
                            new_row.append(val)
                    return new_row

                ws_raw = wb.create_sheet(f"{store_name}原表")
                sheet_groups["raw"].append(ws_raw)
                ws_raw.append(store_headers)
                for raw_row in data.get("raw_rows", []):
                    row_to_write = raw_row + [""] * (len(store_headers) - len(raw_row)) if raw_row else [""] * len(store_headers)
                    ws_raw.append(convert_row_types(row_to_write))

                ws_proc = wb.create_sheet(f"{store_name}处理表")
                sheet_groups["proc"].append(ws_proc)
                proc_headers = store_headers + ["实际商品数量", "总和税运", "商品成本", "订单成本", "综合成本", "平台扣点", "毛利润", "归属日期", "费用项(旧)"] + sorted_fee_names
                ws_proc.append(proc_headers)

                store_agg = {}
                store_link_map = link_map.get(store_name, {})
                store_promo_map = promo_map.get(store_name, {})

                receive_idx = self.find_header_index(store_headers, ["商家实收金额(元)", "商家实收金额"])
                receive_col = get_column_letter(receive_idx + 1) if receive_idx is not None else None

                n_orig = len(store_headers)
                items_count_col = get_column_letter(n_orig + 1)
                tax_col = get_column_letter(n_orig + 2)
                order_cost_col = get_column_letter(n_orig + 4)
                all_cost_col = get_column_letter(n_orig + 5)
                plat_fee_col = get_column_letter(n_orig + 6)

                for row_dict in data["rows"]:
                    orig_row = convert_row_types(row_dict["original_row"])
                    task_id = row_dict["task_id"]

                    if not task_id or task_id not in data["results"]:
                        ws_proc.append(orig_row + [""] * (len(proc_headers) - len(orig_row)))
                        continue

                    res = data["results"][task_id]
                    fee_str = ",".join(f"{k}({v})" for k, v in res.get("fee_details", {}).items())

                    curr_r = ws_proc.max_row + 1
                    subsidy_val = res.get("subsidy", 0.0)

                    row_date_val = task_id.split("_")[0] if task_id else ""

                    sku_key = res.get("sku", "")
                    extracted_qty = 1
                    if "-" in sku_key:
                        suffix = sku_key.split("-")[-1]
                        if suffix.isdigit():
                            extracted_qty = int(suffix)
                    items_count = extracted_qty * res.get("qty", 1)

                    new_cols = [
                        items_count,
                        res.get("query_result", 0.0),
                        res.get("product_cost", 0.0),
                        res.get("order_cost", 0.0),
                        f"={tax_col}{curr_r}+{order_cost_col}{curr_r}",
                        f"={receive_col}{curr_r}*0.006" if receive_col else res.get("platform_fee", 0.0),
                        f"={receive_col}{curr_r}-{all_cost_col}{curr_r}-{plat_fee_col}{curr_r}+{subsidy_val}" if receive_col else res.get("profit", 0.0),
                        row_date_val,
                        fee_str
                    ]
                    for fn in sorted_fee_names:
                        new_cols.append(res.get("fee_details", {}).get(fn, 0.0))

                    ws_proc.append(orig_row + new_cols)

                    curr_row_idx = ws_proc.max_row
                    fee_issue = not res.get("fee_ok", True) and not res.get("is_special", False)
                    if fee_issue or not res.get("cost_ok", True) or not res.get("tax_ok", True):
                        row_fill = red_fill
                    elif res.get("is_special", False):
                        row_fill = pink_fill
                    else:
                        row_fill = white_fill

                    for col_idx in range(1, len(proc_headers) + 1):
                        cell = ws_proc.cell(row=curr_row_idx, column=col_idx)
                        cell.fill = row_fill
                        if row_fill == white_fill:
                            cell.border = gray_border

                    p_id = res.get("product_id", "")
                    p_name = store_link_map.get(p_id, "未知产品")

                    if p_name not in store_agg:
                        store_agg[p_name] = {}
                    if p_id not in store_agg[p_name]:
                        store_agg[p_name][p_id] = {
                            "sales": 0.0, "orders": 0, "items": 0, "promo": get_promo(store_name, p_id),
                            "tax": 0.0, "discount": 0.0, "subsidy": 0.0, "all_cost": 0.0, "platform_fee": 0.0, "profit_base": 0.0
                        }

                    store_agg[p_name][p_id]["sales"] += res.get("receive", 0.0)
                    store_agg[p_name][p_id]["orders"] += 1
                    store_agg[p_name][p_id]["items"] += items_count
                    store_agg[p_name][p_id]["tax"] += res.get("query_result", 0.0)
                    store_agg[p_name][p_id]["discount"] += res.get("discount", 0.0)
                    store_agg[p_name][p_id]["subsidy"] += res.get("subsidy", 0.0)
                    store_agg[p_name][p_id]["all_cost"] += res.get("all_cost", 0.0)
                    store_agg[p_name][p_id]["platform_fee"] += res.get("platform_fee", 0.0)
                    store_agg[p_name][p_id]["profit_base"] += res.get("profit", 0.0)

                    if p_name not in global_agg:
                        global_agg[p_name] = {}
                    if p_id not in global_agg[p_name]:
                        global_agg[p_name][p_id] = {
                            "sales": 0.0, "orders": 0, "items": 0, "promo": get_promo(store_name, p_id),
                            "tax": 0.0, "discount": 0.0, "subsidy": 0.0, "all_cost": 0.0, "platform_fee": 0.0, "profit_base": 0.0
                        }
                    global_agg[p_name][p_id]["sales"] += res.get("receive", 0.0)
                    global_agg[p_name][p_id]["orders"] += 1
                    global_agg[p_name][p_id]["items"] += items_count
                    global_agg[p_name][p_id]["tax"] += res.get("query_result", 0.0)
                    global_agg[p_name][p_id]["discount"] += res.get("discount", 0.0)
                    global_agg[p_name][p_id]["subsidy"] += res.get("subsidy", 0.0)
                    global_agg[p_name][p_id]["all_cost"] += res.get("all_cost", 0.0)
                    global_agg[p_name][p_id]["platform_fee"] += res.get("platform_fee", 0.0)
                    global_agg[p_name][p_id]["profit_base"] += res.get("profit", 0.0)

                global_store_totals[store_name] = {
                    "sales": sum(v["sales"] for links in store_agg.values() for v in links.values()),
                    "orders": sum(v["orders"] for links in store_agg.values() for v in links.values()),
                    "promo": sum(v["promo"] for links in store_agg.values() for v in links.values()),
                    "discount": sum(v["discount"] for links in store_agg.values() for v in links.values()),
                    "subsidy": sum(v["subsidy"] for links in store_agg.values() for v in links.values()),
                    "tax": sum(v["tax"] for links in store_agg.values() for v in links.values()),
                    "all_cost": sum(v["all_cost"] for links in store_agg.values() for v in links.values()),
                    "platform_fee": sum(v["platform_fee"] for links in store_agg.values() for v in links.values()),
                    "profit_base": sum(v["profit_base"] for links in store_agg.values() for v in links.values()),
                }
                global_store_totals[store_name]["profit"] = global_store_totals[store_name]["profit_base"] - global_store_totals[store_name]["promo"]

                store_agg_map[store_name] = store_agg

            sorted_global_agg = sorted(global_agg.items(), key=lambda x: self.get_sort_key(x[0], order_list))

            # Phase 2: 明细表和汇总表
            for store_name, data in all_store_data.items():
                store_agg = store_agg_map[store_name]
                store_promo_map = promo_map.get(store_name, {})

                ws_promo = wb.create_sheet(f"{store_name}推广表")
                sheet_groups["proc"].append(ws_promo)

                if "至" in date_str:
                    ws_promo.append(["商品ID", "日期", "推广花费"])
                    for p_id, p_data in store_promo_map.items():
                        for d in included_dates:
                            daily_val = p_data.get("daily", {}).get(d, 0.0)
                            if daily_val > 0:
                                ws_promo.append([p_id, d, round(daily_val, 2)])
                else:
                    ws_promo.append(["商品ID", "推广花费"])
                    for p_id, promo_val in store_promo_map.items():
                        daily_val = get_promo(store_name, p_id)
                        if daily_val > 0:
                            ws_promo.append([p_id, round(daily_val, 2)])

                for row_idx in range(1, ws_promo.max_row + 1):
                    for col_idx in range(1, ws_promo.max_column + 1):
                        cell = ws_promo.cell(row=row_idx, column=col_idx)
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                        if row_idx == 1:
                            cell.fill = PatternFill("solid", fgColor="D9EAD3")
                        else:
                            cell.fill = white_fill
                        cell.border = gray_border

                ws_detail = wb.create_sheet(f"{store_name}明细表")
                sheet_groups["detail"].append(ws_detail)
                ws_detail.append(["店铺名称", "产品名称", "商品id", "销售额", "订单数量", "推广花费", "平台优惠", "百亿官补", "税运", "平台扣点", "综合成本", "毛利", "毛利率",
                                  "销售额汇总", "订单数量汇总", "推广汇总", "平台优惠汇总", "百亿官补汇总", "税运汇总", "平台扣点汇总", "综合成本汇总", "毛利汇总", "毛利率汇总"])

                align_left_center = Alignment(horizontal="left", vertical="center")

                store_headers = all_store_data[store_name]["headers"]

                def get_proc_col(name_list):
                    idx = self.find_header_index(store_headers, name_list)
                    if idx is not None:
                        return get_column_letter(idx + 1)
                    return None

                proc_sheet_name = f"'{store_name}处理表'"

                id_col_proc = get_proc_col(["商品id", "商品ID", "链接id"])
                receive_col_proc = get_proc_col(["商家实收金额(元)", "商家实收金额"])
                discount_col_proc = get_proc_col(["平台优惠折扣(元)", "平台优惠折扣", "平台优惠"])

                n_orig = len(store_headers)
                tax_col_proc = get_column_letter(n_orig + 2)
                all_cost_col_proc = get_column_letter(n_orig + 5)
                plat_fee_col_proc = get_column_letter(n_orig + 6)

                for p_name, global_links in sorted_global_agg:
                    start_row = ws_detail.max_row + 1
                    store_links = store_agg.get(p_name, {})
                    product_has_data = len(store_links) > 0

                    sorted_p_ids = sorted(global_links.keys())
                    end_row = start_row + len(sorted_p_ids) - 1

                    for idx, p_id in enumerate(sorted_p_ids):
                        curr_r = start_row + idx
                        if product_has_data:
                            right_side = [
                                f"=SUM(D{start_row}:D{end_row})",
                                f"=SUM(E{start_row}:E{end_row})",
                                f"=SUM(F{start_row}:F{end_row})",
                                f"=SUM(G{start_row}:G{end_row})",
                                f"=SUM(H{start_row}:H{end_row})",
                                f"=SUM(I{start_row}:I{end_row})",
                                f"=SUM(J{start_row}:J{end_row})",
                                f"=SUM(K{start_row}:K{end_row})",
                                f"=N{start_row}-U{start_row}-T{start_row}+R{start_row}-P{start_row}",
                                f'=IF(N{start_row}>0, V{start_row}/N{start_row}, 0)'
                            ]
                        else:
                            right_side = [""] * 10

                        if p_id in store_links:
                            stats = store_links[p_id]

                            if id_col_proc and receive_col_proc:
                                sales_formula = f"=SUMIFS({proc_sheet_name}!{receive_col_proc}:{receive_col_proc}, {proc_sheet_name}!{id_col_proc}:{id_col_proc}, \"{p_id}\")"
                                orders_formula = f"=COUNTIFS({proc_sheet_name}!{id_col_proc}:{id_col_proc}, \"{p_id}\")"
                                discount_formula = f"=SUMIFS({proc_sheet_name}!{discount_col_proc}:{discount_col_proc}, {proc_sheet_name}!{id_col_proc}:{id_col_proc}, \"{p_id}\")" if discount_col_proc else round(stats["discount"], 2)
                                tax_formula = f"=SUMIFS({proc_sheet_name}!{tax_col_proc}:{tax_col_proc}, {proc_sheet_name}!{id_col_proc}:{id_col_proc}, \"{p_id}\")"
                                plat_fee_formula = f"=SUMIFS({proc_sheet_name}!{plat_fee_col_proc}:{plat_fee_col_proc}, {proc_sheet_name}!{id_col_proc}:{id_col_proc}, \"{p_id}\")"
                                all_cost_formula = f"=SUMIFS({proc_sheet_name}!{all_cost_col_proc}:{all_cost_col_proc}, {proc_sheet_name}!{id_col_proc}:{id_col_proc}, \"{p_id}\")"
                            else:
                                sales_formula = round(stats["sales"], 2)
                                orders_formula = stats["orders"]
                                discount_formula = round(stats["discount"], 2)
                                tax_formula = round(stats["tax"], 2)
                                plat_fee_formula = round(stats["platform_fee"], 2)
                                all_cost_formula = round(stats["all_cost"], 2)

                            promo_val_col = "C" if "至" in date_str else "B"
                            promo_formula = f"=SUMIFS('{store_name}推广表'!{promo_val_col}:{promo_val_col}, '{store_name}推广表'!A:A, \"{p_id}\")"

                            ws_detail.append([
                                store_name, p_name, p_id,
                                sales_formula, orders_formula, promo_formula, discount_formula,
                                round(stats["subsidy"], 2), tax_formula, plat_fee_formula, all_cost_formula,
                                f"=D{curr_r}-K{curr_r}-J{curr_r}+H{curr_r}-F{curr_r}",
                                f'=IF(D{curr_r}>0, L{curr_r}/D{curr_r}, 0)'
                            ] + right_side)
                        else:
                            ws_detail.append([
                                store_name, p_name, "",
                                "", "", "", "", "", "", "", "", "", ""
                            ] + right_side)

                    if end_row > start_row:
                        ws_detail.merge_cells(start_row=start_row, end_row=end_row, start_column=1, end_column=1)
                        ws_detail.merge_cells(start_row=start_row, end_row=end_row, start_column=2, end_column=2)
                        for col_idx in range(14, 24):
                            ws_detail.merge_cells(start_row=start_row, end_row=end_row, start_column=col_idx, end_column=col_idx)

                    for row_idx, p_id in enumerate(sorted_p_ids, start=start_row):
                        is_empty = p_id not in store_links
                        for col_idx in range(1, 24):
                            cell = ws_detail.cell(row=row_idx, column=col_idx)
                            cell.alignment = align_left_center

                            if col_idx in [1, 2] or col_idx >= 14:
                                cell.fill = yellow_fill if product_has_data else gray_fill
                            else:
                                cell.fill = gray_fill if is_empty else yellow_fill

                            cell.border = gray_border

                            if col_idx in [13, 23]:
                                cell.number_format = '0.00%'

                    ws_detail.append([""] * 23)

            # 全店汇总表
            ws_global = wb.create_sheet("全店汇总表")
            ws_global.append(["产品名称", "商品id", "销售额", "订单数量", "推广花费", "平台优惠", "百亿官补", "税运", "平台扣点", "综合成本", "毛利", "毛利率",
                              "销售额汇总", "订单数量汇总", "推广花费汇总", "平台优惠汇总", "百亿官补汇总", "税运汇总", "平台扣点汇总", "综合成本汇总", "毛利汇总", "毛利率汇总"])
            align_left_center = Alignment(horizontal="left", vertical="center")

            sorted_global_agg = sorted(global_agg.items(), key=lambda x: self.get_sort_key(x[0], order_list))

            for p_name, links in sorted_global_agg:
                start_row = ws_global.max_row + 1
                sorted_p_ids = sorted(links.keys())
                end_row = start_row + len(sorted_p_ids) - 1

                for idx, p_id in enumerate(sorted_p_ids):
                    curr_r = start_row + idx
                    stats = links[p_id]

                    right_side = [
                        f"=SUM(C{start_row}:C{end_row})",
                        f"=SUM(D{start_row}:D{end_row})",
                        f"=SUM(E{start_row}:E{end_row})",
                        f"=SUM(F{start_row}:F{end_row})",
                        f"=SUM(G{start_row}:G{end_row})",
                        f"=SUM(H{start_row}:H{end_row})",
                        f"=SUM(I{start_row}:I{end_row})",
                        f"=SUM(J{start_row}:J{end_row})",
                        f"=M{start_row}-T{start_row}-S{start_row}+Q{start_row}-O{start_row}",
                        f"=IF(M{start_row}>0, U{start_row}/M{start_row}, 0)"
                    ]

                    sales_formula_parts = []
                    orders_formula_parts = []
                    discount_formula_parts = []
                    tax_formula_parts = []
                    plat_fee_formula_parts = []
                    all_cost_formula_parts = []
                    promo_formula_parts = []

                    has_formula = False

                    for s_name in all_store_data.keys():
                        s_headers = all_store_data[s_name]["headers"]

                        def get_s_col(name_list):
                            idx = self.find_header_index(s_headers, name_list)
                            if idx is not None:
                                return get_column_letter(idx + 1)
                            return None

                        ps_name = f"'{s_name}处理表'"
                        s_id_col = get_s_col(["商品id", "商品ID", "链接id"])
                        s_rec_col = get_s_col(["商家实收金额(元)", "商家实收金额"])
                        s_disc_col = get_s_col(["平台优惠折扣(元)", "平台优惠折扣", "平台优惠"])

                        s_n_orig = len(s_headers)
                        s_tax_col = get_column_letter(s_n_orig + 2)
                        s_all_cost_col = get_column_letter(s_n_orig + 5)
                        s_plat_fee_col = get_column_letter(s_n_orig + 6)

                        promo_val_col = "C" if "至" in date_str else "B"
                        promo_formula_parts.append(f"SUMIFS('{s_name}推广表'!{promo_val_col}:{promo_val_col}, '{s_name}推广表'!A:A, \"{p_id}\")")

                        if s_id_col and s_rec_col:
                            has_formula = True
                            sales_formula_parts.append(f"SUMIFS({ps_name}!{s_rec_col}:{s_rec_col}, {ps_name}!{s_id_col}:{s_id_col}, \"{p_id}\")")
                            orders_formula_parts.append(f"COUNTIFS({ps_name}!{s_id_col}:{s_id_col}, \"{p_id}\")")
                            if s_disc_col:
                                discount_formula_parts.append(f"SUMIFS({ps_name}!{s_disc_col}:{s_disc_col}, {ps_name}!{s_id_col}:{s_id_col}, \"{p_id}\")")
                            tax_formula_parts.append(f"SUMIFS({ps_name}!{s_tax_col}:{s_tax_col}, {ps_name}!{s_id_col}:{s_id_col}, \"{p_id}\")")
                            plat_fee_formula_parts.append(f"SUMIFS({ps_name}!{s_plat_fee_col}:{s_plat_fee_col}, {ps_name}!{s_id_col}:{s_id_col}, \"{p_id}\")")
                            all_cost_formula_parts.append(f"SUMIFS({ps_name}!{s_all_cost_col}:{s_all_cost_col}, {ps_name}!{s_id_col}:{s_id_col}, \"{p_id}\")")

                    if has_formula:
                        g_sales_val = "=" + "+".join(sales_formula_parts)
                        g_orders_val = "=" + "+".join(orders_formula_parts)
                        g_discount_val = "=" + "+".join(discount_formula_parts) if discount_formula_parts else round(stats["discount"], 2)
                        g_tax_val = "=" + "+".join(tax_formula_parts)
                        g_plat_fee_val = "=" + "+".join(plat_fee_formula_parts)
                        g_all_cost_val = "=" + "+".join(all_cost_formula_parts)
                        g_promo_val = "=" + "+".join(promo_formula_parts) if promo_formula_parts else round(stats["promo"], 2)
                    else:
                        g_sales_val = round(stats["sales"], 2)
                        g_orders_val = stats["orders"]
                        g_discount_val = round(stats["discount"], 2)
                        g_tax_val = round(stats["tax"], 2)
                        g_plat_fee_val = round(stats["platform_fee"], 2)
                        g_all_cost_val = round(stats["all_cost"], 2)
                        g_promo_val = round(stats["promo"], 2)

                    ws_global.append([
                        p_name, p_id,
                        g_sales_val, g_orders_val, g_promo_val, g_discount_val,
                        round(stats["subsidy"], 2), g_tax_val, g_plat_fee_val, g_all_cost_val,
                        f"=C{curr_r}-J{curr_r}-I{curr_r}+G{curr_r}-E{curr_r}",
                        f"=IF(C{curr_r}>0, K{curr_r}/C{curr_r}, 0)"
                    ] + right_side)

                if end_row > start_row:
                    ws_global.merge_cells(start_row=start_row, end_row=end_row, start_column=1, end_column=1)
                    for col_idx in range(13, 23):
                        ws_global.merge_cells(start_row=start_row, end_row=end_row, start_column=col_idx, end_column=col_idx)

                for row_idx in range(start_row, end_row + 1):
                    for col_idx in range(1, 23):
                        cell = ws_global.cell(row=row_idx, column=col_idx)
                        cell.alignment = align_left_center
                        cell.fill = yellow_fill
                        cell.border = gray_border

                        if col_idx in [12, 22]:
                            cell.number_format = '0.00%'

                ws_global.append([""] * 22)

            for _ in range(3):
                ws_global.append([""] * 22)

            # 店铺销售数据汇总
            store_summary_title_row = ws_global.max_row + 1
            ws_global.append(["店铺销售数据汇总"])
            ws_global.merge_cells(start_row=store_summary_title_row, end_row=store_summary_title_row, start_column=1, end_column=11)
            title_cell = ws_global.cell(row=store_summary_title_row, column=1)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            title_cell.font = openpyxl.styles.Font(bold=True, size=12)
            title_cell.fill = PatternFill("solid", fgColor="D9EAD3")
            title_cell.border = gray_border
            for col_idx in range(2, 12):
                ws_global.cell(row=store_summary_title_row, column=col_idx).border = gray_border

            header_row = ws_global.max_row + 1
            ws_global.append(["店铺名称", "销售额", "订单数量", "推广花费", "平台优惠", "百亿官补", "税运", "平台扣点", "综合成本", "毛利", "毛利率"])

            for s_name, s_stats in global_store_totals.items():
                curr_r = ws_global.max_row + 1
                promo_val_col = "C" if "至" in date_str else "B"
                ws_global.append([
                    s_name,
                    round(s_stats["sales"], 2),
                    s_stats["orders"],
                    f"=SUM('{s_name}推广表'!{promo_val_col}:{promo_val_col})",
                    round(s_stats["discount"], 2),
                    round(s_stats["subsidy"], 2),
                    round(s_stats["tax"], 2),
                    round(s_stats["platform_fee"], 2),
                    round(s_stats["all_cost"], 2),
                    f"=B{curr_r}-I{curr_r}-H{curr_r}+F{curr_r}-D{curr_r}",
                    f"=IF(B{curr_r}>0, J{curr_r}/B{curr_r}, 0)"
                ])

            for row_idx in range(header_row, ws_global.max_row + 1):
                for col_idx in range(1, 12):
                    cell = ws_global.cell(row=row_idx, column=col_idx)
                    cell.alignment = align_left_center
                    if row_idx == header_row:
                        cell.fill = PatternFill("solid", fgColor="D9EAD3")
                    else:
                        cell.fill = yellow_fill
                    cell.border = gray_border

                    if col_idx == 11 and row_idx > header_row:
                        cell.number_format = '0.00%'

            # 销售计划登记表
            ws_plan = wb.create_sheet("销售计划登记")
            ws_plan.append(["产品名称", "销售额", "订单数量", "商品数量", "税运", "综合成本", "推广花费", "推广费比", "预估净毛利", "预估净毛利率"])

            for p_name, links in sorted_global_agg:
                plan_sales_parts = []
                plan_orders_parts = []
                plan_items_parts = []
                plan_tax_parts = []
                plan_all_cost_parts = []
                promo_formula_parts = []
                plan_promo = sum(v["promo"] for v in links.values())
                plan_profit_base_parts = []

                has_plan_formula = False

                for s_name in all_store_data.keys():
                    s_headers = all_store_data[s_name]["headers"]

                    def get_s_col(name_list):
                        idx = self.find_header_index(s_headers, name_list)
                        if idx is not None:
                            return get_column_letter(idx + 1)
                        return None

                    ps_name = f"'{s_name}处理表'"
                    s_id_col = get_s_col(["商品id", "商品ID", "链接id"])
                    s_rec_col = get_s_col(["商家实收金额(元)", "商家实收金额"])

                    s_n_orig = len(s_headers)
                    s_items_col = get_column_letter(s_n_orig + 1)
                    s_tax_col = get_column_letter(s_n_orig + 2)
                    s_all_cost_col = get_column_letter(s_n_orig + 5)
                    s_profit_col = get_column_letter(s_n_orig + 7)

                    promo_val_col = "C" if "至" in date_str else "B"

                    # [Web版修复 R2] 原 GUI 此处误用循环外残留的 store_link_map
                    # (最后一个店铺的映射)，导致销售计划登记表只统计最后一个店铺
                    s_link_map = link_map.get(s_name, {})

                    for p_id in links.keys():
                        if p_id not in s_link_map:
                            continue

                        promo_formula_parts.append(f"SUMIFS('{s_name}推广表'!{promo_val_col}:{promo_val_col}, '{s_name}推广表'!A:A, \"{p_id}\")")
                        if s_id_col and s_rec_col:
                            has_plan_formula = True
                            plan_sales_parts.append(f"SUMIFS({ps_name}!{s_rec_col}:{s_rec_col}, {ps_name}!{s_id_col}:{s_id_col}, \"{p_id}\")")
                            plan_orders_parts.append(f"COUNTIFS({ps_name}!{s_id_col}:{s_id_col}, \"{p_id}\")")
                            plan_items_parts.append(f"SUMIFS({ps_name}!{s_items_col}:{s_items_col}, {ps_name}!{s_id_col}:{s_id_col}, \"{p_id}\")")
                            plan_tax_parts.append(f"SUMIFS({ps_name}!{s_tax_col}:{s_tax_col}, {ps_name}!{s_id_col}:{s_id_col}, \"{p_id}\")")
                            plan_all_cost_parts.append(f"SUMIFS({ps_name}!{s_all_cost_col}:{s_all_cost_col}, {ps_name}!{s_id_col}:{s_id_col}, \"{p_id}\")")
                            plan_profit_base_parts.append(f"SUMIFS({ps_name}!{s_profit_col}:{s_profit_col}, {ps_name}!{s_id_col}:{s_id_col}, \"{p_id}\")")

                curr_r = ws_plan.max_row + 1

                if has_plan_formula:
                    plan_sales_val = "=" + "+".join(plan_sales_parts) if plan_sales_parts else "0"
                    plan_orders_val = "=" + "+".join(plan_orders_parts) if plan_orders_parts else "0"
                    plan_items_val = "=" + "+".join(plan_items_parts) if plan_items_parts else "0"
                    plan_tax_val = "=" + "+".join(plan_tax_parts) if plan_tax_parts else "0"
                    plan_all_cost_val = "=" + "+".join(plan_all_cost_parts) if plan_all_cost_parts else "0"
                    plan_promo_val = "=" + "+".join(promo_formula_parts) if promo_formula_parts else round(plan_promo, 2)

                    if plan_profit_base_parts:
                        base_profit_formula = "+".join(plan_profit_base_parts)
                        plan_profit_val = f"=({base_profit_formula})-({plan_promo_val[1:] if str(plan_promo_val).startswith('=') else plan_promo_val})"
                    else:
                        plan_profit_val = f"=0-({plan_promo_val[1:] if str(plan_promo_val).startswith('=') else plan_promo_val})"
                else:
                    plan_sales_val = round(sum(v["sales"] for v in links.values()), 2)
                    plan_orders_val = sum(v["orders"] for v in links.values())
                    plan_items_val = sum(v["items"] for v in links.values())
                    plan_tax_val = round(sum(v["tax"] for v in links.values()), 2)
                    plan_all_cost_val = round(sum(v["all_cost"] for v in links.values()), 2)
                    plan_promo_val = round(plan_promo, 2)
                    plan_profit_val = round(sum(v["profit_base"] for v in links.values()) - plan_promo, 2)

                ws_plan.append([
                    p_name,
                    plan_sales_val,
                    plan_orders_val,
                    plan_items_val,
                    plan_tax_val,
                    plan_all_cost_val,
                    plan_promo_val,
                    f"=IF(B{curr_r}>0, G{curr_r}/B{curr_r}, 0)",
                    plan_profit_val,
                    f"=IF(B{curr_r}>0, I{curr_r}/B{curr_r}, 0)"
                ])

            for row_idx in range(1, ws_plan.max_row + 1):
                for col_idx in range(1, 11):
                    cell = ws_plan.cell(row=row_idx, column=col_idx)
                    cell.alignment = align_left_center
                    if row_idx == 1:
                        cell.fill = PatternFill("solid", fgColor="D9EAD3")
                    else:
                        cell.fill = yellow_fill
                    cell.border = gray_border

                    if col_idx in [8, 10] and row_idx > 1:
                        cell.number_format = '0.00%'

            # 特定产品每日数据子表
            daily_sheets_to_add = []
            target_products = ["dokkan酵素香槟金", "movefree红瓶 200粒"]
            daily_headers = ["日期", "税运", "平台扣点", "订单数量", "商品数量", "销售额", "推广花费", "投产比", "消耗率", "毛利", "毛利率"]

            for target_p in target_products:
                ws_daily = wb.create_sheet(f"{target_p}每日数据")
                daily_sheets_to_add.append(ws_daily)
                ws_daily.append(daily_headers)

                for single_date in sorted(list(all_date_store_data.keys())):
                    if "至" in single_date:
                        continue

                    s_date_data = all_date_store_data[single_date]
                    d_items = 0
                    d_promo = 0.0

                    daily_tax_parts = []
                    daily_plat_fee_parts = []
                    daily_orders_parts = []
                    daily_sales_parts = []
                    daily_items_parts = []
                    daily_profit_base_parts = []
                    daily_promo_parts = []

                    has_daily_formula = False

                    for s_name, s_info in s_date_data.items():
                        s_results = s_info.get("results", {})
                        store_promo_map = promo_map.get(s_name, {})
                        store_link_map = link_map.get(s_name, {})

                        target_p_links = [l_id for l_id, p_name in store_link_map.items() if self.get_sort_key(p_name, [target_p])[0] == 0]

                        s_headers = s_info["headers"]

                        def get_s_col(name_list):
                            idx = self.find_header_index(s_headers, name_list)
                            if idx is not None:
                                return get_column_letter(idx + 1)
                            return None

                        ps_name = f"'{s_name}处理表'"
                        s_id_col = get_s_col(["商品id", "商品ID", "链接id"])
                        s_rec_col = get_s_col(["商家实收金额(元)", "商家实收金额"])

                        s_n_orig = len(s_headers)
                        s_items_col = get_column_letter(s_n_orig + 1)
                        s_tax_col = get_column_letter(s_n_orig + 2)
                        s_plat_fee_col = get_column_letter(s_n_orig + 6)
                        s_profit_col = get_column_letter(s_n_orig + 7)
                        s_date_col = get_column_letter(s_n_orig + 8)

                        for p_id in target_p_links:
                            if "至" in date_str:
                                daily_promo_parts.append(f"SUMIFS('{s_name}推广表'!C:C, '{s_name}推广表'!A:A, \"{p_id}\", '{s_name}推广表'!B:B, \"{single_date}\")")
                            else:
                                daily_promo_parts.append(f"SUMIFS('{s_name}推广表'!B:B, '{s_name}推广表'!A:A, \"{p_id}\")")

                            if s_id_col and s_rec_col:
                                has_daily_formula = True
                                daily_sales_parts.append(f"SUMIFS({ps_name}!{s_rec_col}:{s_rec_col}, {ps_name}!{s_id_col}:{s_id_col}, \"{p_id}\", {ps_name}!{s_date_col}:{s_date_col}, \"{single_date}\")")
                                daily_orders_parts.append(f"COUNTIFS({ps_name}!{s_id_col}:{s_id_col}, \"{p_id}\", {ps_name}!{s_date_col}:{s_date_col}, \"{single_date}\")")
                                daily_items_parts.append(f"SUMIFS({ps_name}!{s_items_col}:{s_items_col}, {ps_name}!{s_id_col}:{s_id_col}, \"{p_id}\", {ps_name}!{s_date_col}:{s_date_col}, \"{single_date}\")")
                                daily_tax_parts.append(f"SUMIFS({ps_name}!{s_tax_col}:{s_tax_col}, {ps_name}!{s_id_col}:{s_id_col}, \"{p_id}\", {ps_name}!{s_date_col}:{s_date_col}, \"{single_date}\")")
                                daily_plat_fee_parts.append(f"SUMIFS({ps_name}!{s_plat_fee_col}:{s_plat_fee_col}, {ps_name}!{s_id_col}:{s_id_col}, \"{p_id}\", {ps_name}!{s_date_col}:{s_date_col}, \"{single_date}\")")
                                daily_profit_base_parts.append(f"SUMIFS({ps_name}!{s_profit_col}:{s_profit_col}, {ps_name}!{s_id_col}:{s_id_col}, \"{p_id}\", {ps_name}!{s_date_col}:{s_date_col}, \"{single_date}\")")

                        for task_id, res in s_results.items():
                            p_id = res.get("product_id", "")
                            if p_id in target_p_links:
                                sku_key = res.get("sku", "")
                                extracted_qty = 1
                                if "-" in sku_key:
                                    suffix = sku_key.split("-")[-1]
                                    if suffix.isdigit():
                                        extracted_qty = int(suffix)
                                items_count = extracted_qty * res.get("qty", 1)
                                d_items += items_count

                        for p_id in target_p_links:
                            d_promo += store_promo_map.get(p_id, {}).get("daily", {}).get(single_date, 0.0)

                    curr_r = ws_daily.max_row + 1

                    if has_daily_formula:
                        d_tax_val = "=" + "+".join(daily_tax_parts) if daily_tax_parts else "0"
                        d_plat_fee_val = "=" + "+".join(daily_plat_fee_parts) if daily_plat_fee_parts else "0"
                        d_orders_val = "=" + "+".join(daily_orders_parts) if daily_orders_parts else "0"
                        d_items_val = "=" + "+".join(daily_items_parts) if daily_items_parts else "0"
                        d_sales_val = "=" + "+".join(daily_sales_parts) if daily_sales_parts else "0"
                        d_promo_val = "=" + "+".join(daily_promo_parts) if daily_promo_parts else round(d_promo, 2)

                        if daily_profit_base_parts:
                            base_profit_f = "+".join(daily_profit_base_parts)
                            d_profit_val = f"=({base_profit_f})-({d_promo_val[1:] if str(d_promo_val).startswith('=') else d_promo_val})"
                        else:
                            d_profit_val = f"=0-({d_promo_val[1:] if str(d_promo_val).startswith('=') else d_promo_val})"
                    else:
                        d_tax_val = d_plat_fee_val = d_orders_val = d_sales_val = d_profit_val = 0
                        d_items_val = d_items
                        d_promo_val = round(d_promo, 2)

                    ws_daily.append([
                        single_date,
                        d_tax_val,
                        d_plat_fee_val,
                        d_orders_val,
                        d_items_val,
                        d_sales_val,
                        d_promo_val,
                        f"=IF(G{curr_r}>0, F{curr_r}/G{curr_r}, 0)",
                        f"=IF(F{curr_r}>0, G{curr_r}/F{curr_r}, 0)",
                        d_profit_val,
                        f"=IF(F{curr_r}>0, J{curr_r}/F{curr_r}, 0)"
                    ])

                for row_idx in range(1, ws_daily.max_row + 1):
                    for col_idx in range(1, 12):
                        cell = ws_daily.cell(row=row_idx, column=col_idx)
                        cell.alignment = align_left_center
                        if row_idx == 1:
                            cell.fill = PatternFill("solid", fgColor="D9EAD3")
                        else:
                            cell.fill = yellow_fill
                        cell.border = gray_border

                        if col_idx in [9, 11] and row_idx > 1:
                            cell.number_format = '0.00%'
                        elif col_idx == 8 and row_idx > 1:
                            cell.number_format = '0.00'

            # 自适应列宽
            sheets_for_width = [ws_global, ws_plan] + daily_sheets_to_add
            for sheet_group in [sheet_groups["detail"], sheets_for_width]:
                for ws in sheet_group:
                    for col in ws.columns:
                        max_length = 0
                        col_letter = col[0].column_letter
                        for cell in col:
                            if cell.value:
                                val_str = str(cell.value)
                                length = sum(1.8 if ord(c) > 127 else 1.1 for c in val_str)
                                if length > max_length:
                                    max_length = length
                        adjusted_width = min(max(max_length + 2, 10), 50)
                        ws.column_dimensions[col_letter].width = adjusted_width

            # 重排 Sheet 顺序（替代 GUI 的 wb._sheets 私有属性赋值）
            ordered_sheets = []
            ordered_sheets.extend(sheet_groups["raw"])
            ordered_sheets.extend(sheet_groups["proc"])
            ordered_sheets.extend(sheet_groups["detail"])
            ordered_sheets.append(ws_global)
            ordered_sheets.append(ws_plan)
            ordered_sheets.extend(daily_sheets_to_add)

            for i, ws in enumerate(ordered_sheets):
                current = wb.sheetnames.index(ws.title)
                if current != i:
                    wb.move_sheet(ws, offset=i - current)

            # 保存文件
            output_filename = os.path.join(output_dir, f"{date_str}全店铺销售数据汇总表.xlsx")
            wb.save(output_filename)
            result_files.append(os.path.abspath(output_filename))
            self.log(f"已生成: {output_filename}")

        self.progress(100, "处理完成")
        self.log("任务完成！所有汇总文件已生成。")

        return {
            "files": result_files,
            "stats": {
                "raw_rows": raw_total,
                "total_tasks": total_tasks,
                "queried": processed_count,
                "red_rows": None,
                "cache_size": len(self.tax_cache),
            },
            "tax_cache": self.tax_cache,
        }
