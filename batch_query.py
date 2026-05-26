import requests
import json
import openpyxl
import time
import os

# ===================== 配置部分 =====================
# 请在这里替换成你自己的cookie字符串 (从test_python.py复制)
COOKIE = r"xlly_s=1; cn_account_lang=zh_CN; x-hng=lang=zh-CN; isg=BOfnyVnGQlsoU8bXhJWnuqHydhuxbLtOfzp4WrlVlXadqA5qxjounEvrzqg2QJPG; lvt=1772432731829; cncc=fb109569514aa968c18a30bd9ccb909d; TwAhx8HL=C6BCB215B673113B88DEFEF22D0D5CD441D3D197ABA710AE36592A8446521ABF704A4A73694B4970B1A24909F661D091A53A1457C557B028E643F3FD6A9648E39F4237DD45D2D7D16297F502326F30DB3CB04FD8ED746B1B2E03282863AE59E817CDC3561B36201385CC8C2E0D9677A430C4FDB8FCA05F8E369E74A334C708D5B783B140A4856B58E704E47C67746EB42FDC758315A88B1ABADC4C2A0B142FA7C70E3191AD4A1A1884DEEF92CB4FFBA93C7F6A1F3A9E6EA26CF8575C634CF36D831C4DBCA383788679FA1670FCF949506112BD4B511581AC7CC293AD49C2F4BAB2598908D4FBE9F4CBD43D8A63686A98; SL34syaT=8C5D983BCBE6B9EFC7610EFF8ADB68A1; accountId=MTgxNjc3ODEwMzIyMw==; cacbi=MTgxNjc3ODEwMzIyMw==; cpcode=4399135568313; isLogin=true; account=Y24t6Laj5ZGz54y05rW35aSW5peX6Iiw5bqXOui0pOW+tw==; userLoginType=cic; cn-gateway-useagent=pc; bizUserId=2212853812821; cna=XRssIkF/Px4CAbcRfm9ABbQy; CURRENT_LANGUAGE=zh; x-hng=lang=zh-CN&domain=b.cainiao.com; cf_counter_nick=%E8%B6%A3%E5%91%B3%E7%8C%B4%E6%B5%B7%E5%A4%96%E6%97%97%E8%88%B0%E5%BA%97; loginType=havanaTokenSSO-TAOBAO; cnUId=1816778103223; tfstk=gjztotYXh9XGJF_CeFjhizhw4n5hZMVadRPWoxDMcJeLHv4G_lVckqeIKPDgsVMvDkN4jq2c_mwLhR1ZiPfaJ-FYFlcm5ORxkSezjcgimjLYMvhMj1oclCzLeEYGQGPXkq0fETblr5lZuq6oO0U5JdljGf_mfhgI7jY_OwScr5PwOYuAnh_lkUH3ejMbhcGIAjkI1jM_h2gIavTsGFGXd6hqdqTsCAtIObhkChibhW1KGvMjlcwjR6hqdxgjlb5dHvOsnEEphqL3inKXkEUK6cHvovYxvP8omY1tQF8ZBfn9j5MplEaLc6gt9AsytxuiQ5GT3a8Z5mEbS2a1dZg74JUxV4Ic_cNYv-ugRstsFl2imPopChG-WAn75DbROWibvy3g5t-ztWHs22qhYOmmWRErUcsFL-FKI-atAdBnnuV35De5KeeqDoaiRP6Cyg-zrzI2fLD-ih1d9n-qfXrzAC2cd0UvCXHl6ZK20D5E9Yfd9n-qfXlKE1B90noFT"

# 接口地址
API_URL = "https://b.cainiao.com/merchantchargeorder/queryChargeOrderListNew"

def get_charge_order_list(order_no):
    """
    发送POST请求获取菜鸟充值订单列表
    """
    # 模拟浏览器的请求头
    headers = {
        "accept": "application/json, text/plain, */*",
        "bx-v": "2.5.36",
        "content-type": "application/json",
        "h-csrf": "635e77ca-003a-498a-9a58-18c213306b08",
        "referer": "https://b.cainiao.com/cf_seller/charge-order/charge-order-query",
        "sec-ch-ua": '"Not:A-Brand";v="99", "Microsoft Edge";v="145", "Chromium";v="145"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/145.0.0.0 Safari/537.36 Edg/145.0.0.0",
        "Cookie": COOKIE
    }

    # POST请求参数
    post_data = {
        "orderNo": order_no,
        "currentPage": 1,
        "pageSize": 10,
        "orderNoType": 0
    }

    try:
        # 发送POST请求
        response = requests.post(
            url=API_URL,
            headers=headers,
            data=json.dumps(post_data),
            timeout=30,
            allow_redirects=False
        )

        # 检查是否发生重定向
        if response.status_code in (301, 302, 303, 307, 308):
            print(f"错误：请求被重定向到 {response.headers.get('Location')} (可能是Cookie失效)")
            return None

        # 检查响应状态码
        if response.status_code != 200:
            print(f"请求失败: 状态码 {response.status_code}")
            return None

        return response.json()

    except Exception as e:
        print(f"请求异常: {e}")
        return None

def process_excel(filename):
    if not os.path.exists(filename):
        print(f"文件不存在: {filename}")
        return

    print(f"开始处理文件: {filename}")
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        
        # 假设第一行是表头，从第二行开始
        # openpyxl 行和列都是从1开始
        # U列是第21列
        # X列是第24列
        
        row_count = 0
        success_count = 0
        
        for row in ws.iter_rows(min_row=2):
            row_count += 1
            # 获取U列单元格 (索引20)
            order_no_cell = row[20] 
            order_no = order_no_cell.value
            
            if not order_no:
                continue
                
            order_no_str = str(order_no).strip()
            print(f"正在处理第 {row_count} 行，单号: {order_no_str}")
            
            result = get_charge_order_list(order_no_str)
            
            total_quoted_amount = 0.0
            if result and result.get("success") and result.get("data"):
                data_list = result.get("data")
                if isinstance(data_list, list):
                    for item in data_list:
                        amount = item.get("quotedAmount")
                        if amount:
                            total_quoted_amount += float(amount)
                    success_count += 1
            else:
                 print(f"  查询无数据或失败")
            
            # X列是第24列 (索引23)
            # 确保行有足够的单元格，如果没有，则创建
            if len(row) <= 23:
                # 如果当前行单元格不足，我们需要通过 worksheet.cell 来访问并写入
                # openpyxl 行号是从1开始，列号是从1开始
                # row_count 已经是相对 min_row=2 的计数了，我们需要绝对行号
                # 但 iter_rows 返回的是元组，我们也可以通过 cell 对象获取行号
                current_row_idx = row[0].row
                target_cell = ws.cell(row=current_row_idx, column=24)
            else:
                target_cell = row[23]
                
            target_cell.value = total_quoted_amount
            print(f"  总金额: {total_quoted_amount}")
            
            # 避免请求过快，随机休眠
            time.sleep(0.2)
            
        output_filename = filename.replace(".xlsx", "_result.xlsx")
        wb.save(output_filename)
        print(f"\n处理完成！共处理 {row_count} 行，成功获取 {success_count} 个。")
        print(f"结果已保存为: {output_filename}")
        
    except Exception as e:
        print(f"处理Excel失败: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # 这里指定要处理的文件名
    process_excel("1.xlsx")
