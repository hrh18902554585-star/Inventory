# -*- coding: utf-8 -*-
"""任务接口 API 集成测试（pytest + Flask test_client）

覆盖：
1. POST /api/jobs 缺 cookie → code 2/3 类错误
2. 正常提交 → task_id，状态 pending/running → 轮询至 done
3. done 后 result_files 非空；下载 200 且内容非空（xlsx 的 zip 魔数 PK）
4. 日志增量：since=len 空 / since=0 全量
5. GET /api/jobs 列表包含该任务
6. POST rerun → 新 task_id
7. DELETE → 删除后 GET 404/code=4
8. 税运接口失败（{"error": "请求失败"}）→ 任务仍 done，标红行，不报错
"""
import io
import json
import os
import sys
import time
import types

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "profit_service"))

import openpyxl

import configs
import db
import engine
import files_store
import security
import stats
import task_manager
import tax_cache

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

FAKE_API_OK = {
    "success": True,
    "data": [
        {"quotedAmount": 5.5, "feeName": "基础服务费"},
        {"quotedAmount": 3.0, "feeName": "进口关税"},
    ],
}
FAKE_API_FAIL = {"error": "请求失败"}


# ---------- fixture 生成 ----------
def make_xlsx_bytes(sheet_name, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for r in rows:
        ws.append(r)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


COST_BYTES = make_xlsx_bytes(
    "sku商品编码",
    ["组合编码", "单品编码", "商品名称", "数量", "单品成本", "sku成本"],
    [["SKU-A", "", "雅培金装", 1, 50, 50], ["SKU-B", "", "雅培金装", 1, 60, 60]],
)
LINK_BYTES = make_xlsx_bytes("产品链接汇总表", ["店铺名称", "产品名称", "商品id"], [["庆余", "雅培金装", "1001"]])
PROMO_BYTES = make_xlsx_bytes("推广报表", ["商品ID", "总花费(元)"], [["1001", 20]])
SUBSIDY_BYTES = make_xlsx_bytes("官补映射表", ["链接id", "规格编码", "官补金额"], [["1001", "SKU-A", 10]])


def make_order_csv_bytes():
    out = io.StringIO()
    import csv
    w = csv.writer(out)
    w.writerow(["订单号", "订单状态", "商品id", "商家编码-规格维度", "商品数量(件)", "商家实收金额(元)", "平台优惠折扣(元)", "支付时间"])
    w.writerow(["20260001", "已发货，待收货", "1001", "SKU-A", 2, 200, 10, "2026-08-15 10:00:00"])
    w.writerow(["20260002", "已收货", "1001", "SKU-B", 1, 150, 5, "2026-08-15 11:00:00"])
    w.writerow(["20260003", "已退款", "1001", "SKU-A", 1, 100, 0, "2026-08-15 12:00:00"])
    return out.getvalue().encode("utf-8-sig")


ORDER_CSV_BYTES = make_order_csv_bytes()


# ---------- 共享环境 ----------
@pytest.fixture(autouse=True)
def env(tmp_path, monkeypatch):
    """目录常量 → tmp_path；初始化 db/缓存/统计/任务调度；返回 (client, headers)"""
    monkeypatch.setattr(db, "DB_PATH", str(tmp_path / "app.db"))
    monkeypatch.setattr(security, "TOKEN_FILE", str(tmp_path / "token.txt"))
    monkeypatch.setattr(configs, "CONFIGS_DIR", str(tmp_path / "configs"))
    monkeypatch.setattr(files_store, "UPLOADS_DIR", str(tmp_path / "uploads"))
    monkeypatch.setattr(task_manager, "TASKS_DIR", str(tmp_path / "tasks"))
    monkeypatch.setattr(task_manager, "OUTPUTS_DIR", str(tmp_path / "outputs"))
    monkeypatch.setattr(engine, "get_charge_order_list", lambda order_no, cookie: FAKE_API_OK)

    import app as app_mod
    monkeypatch.setattr(app_mod, "COOKIE_FILE", str(tmp_path / "cookie.txt"))

    db.init_db(str(tmp_path / "app.db"))
    tax_cache.load_cache(str(tmp_path / "tax_cache.json"))
    stats.load_stats(str(tmp_path / "stats.json"))
    task_manager.init_task_manager()

    client = app_mod.create_app().test_client()
    token = security.get_token()
    headers = {"X-Api-Token": token}
    return client, headers


def upload_config(client, headers, cfg_type, filename, data):
    r = client.post("/api/configs/upload", data={"type": cfg_type, "file": (io.BytesIO(data), filename)},
                    headers=headers, content_type="multipart/form-data")
    assert r.status_code == 200, r.data
    body = r.get_json()
    assert body["code"] == 0, body
    return body["data"]["uploaded"][0]


@pytest.fixture()
def full_setup(env):
    """上传成本/链接/推广/官补配置 + 订单文件，返回可复用 ids"""
    client, headers = env
    ids = {
        "cost": upload_config(client, headers, "cost", "成本表.xlsx", COST_BYTES),
        "link": upload_config(client, headers, "link", "产品链接汇总表.xlsx", LINK_BYTES),
        "promo1": upload_config(client, headers, "promo", "推广报表_庆余.xlsx", PROMO_BYTES),
        "subsidy": upload_config(client, headers, "subsidy", "官补映射表.xlsx", SUBSIDY_BYTES),
    }
    r = client.post("/api/uploads", data={"file": (io.BytesIO(ORDER_CSV_BYTES), "庆余订单_20260815.csv")},
                    headers=headers, content_type="multipart/form-data")
    assert r.status_code == 200, r.data
    body = r.get_json()
    assert body["code"] == 0, body
    ids["order_file"] = body["data"]["uploaded"][0]["id"]
    return ids


def make_job_params(ids, **overrides):
    params = {
        "order_files": [ids["order_file"]],
        "configs": {
            "cost": ids["cost"]["version"],
            "link": ids["link"]["version"],
            "promo": [ids["promo1"]["version"]],
            "subsidy": ids["subsidy"]["version"],
        },
        "cookie": "fake_cookie_abc123",
        "thread_count": 2,
        "enable_tax": True,
        "exclude_ids": "",
    }
    params.update(overrides)
    return params


def submit_job(client, headers, ids, **overrides):
    r = client.post("/api/jobs", json=make_job_params(ids, **overrides), headers=headers)
    assert r.status_code == 200, r.data
    body = r.get_json()
    assert body["code"] == 0, body
    return body["data"]["task_id"]


def wait_done(client, headers, task_id, timeout=30, interval=0.2):
    deadline = time.time() + timeout
    data = None
    while time.time() < deadline:
        r = client.get(f"/api/jobs/{task_id}?since=0", headers=headers)
        assert r.status_code == 200, r.data
        data = r.get_json()["data"]
        if data["status"] in ("done", "error"):
            return data
        time.sleep(interval)
    raise AssertionError(f"任务 {task_id} 超时未完成: {data}")


def download_content(client, headers, task_id, filename):
    r = client.get(f"/api/jobs/{task_id}/download/{filename}", headers=headers)
    assert r.status_code == 200, r.data
    content = r.data
    assert content, "下载内容为空"
    return content


# ---------- 用例 ----------
def test_submit_missing_cookie(env, full_setup):
    """1. 缺 cookie → code 2/3 类错误"""
    client, headers = env
    r = client.post("/api/jobs", json=make_job_params(full_setup, cookie=""), headers=headers)
    assert r.status_code == 200
    body = r.get_json()
    assert body["code"] in (2, 3), body
    assert body["message"]


def test_submit_and_poll_done(env, full_setup):
    """2. 正常提交 → task_id，pending/running → 轮询至 done"""
    client, headers = env
    task_id = submit_job(client, headers, full_setup)
    assert task_id.startswith("j_")
    r = client.get(f"/api/jobs/{task_id}?since=0", headers=headers)
    assert r.status_code == 200
    init = r.get_json()["data"]
    assert init["status"] in ("pending", "running"), init
    data = wait_done(client, headers, task_id)
    assert data["status"] == "done", data
    assert data["error"] == ""
    assert data["progress"] == 100


def test_result_files_and_download(env, full_setup):
    """3. done 后 result_files 非空；下载 200 且内容非空（zip 魔数 PK）"""
    client, headers = env
    task_id = submit_job(client, headers, full_setup)
    data = wait_done(client, headers, task_id)
    assert data["status"] == "done"
    assert data["result_files"], "result_files 为空"
    for fname in data["result_files"]:
        content = download_content(client, headers, task_id, fname)
        assert content[:2] == b"PK", "xlsx 应为 zip 魔数 PK"


def test_log_increment(env, full_setup):
    """4. 日志增量：since=len 返回空，since=0 返回全量"""
    client, headers = env
    task_id = submit_job(client, headers, full_setup)
    wait_done(client, headers, task_id)
    r = client.get(f"/api/jobs/{task_id}?since=0", headers=headers)
    assert r.status_code == 200
    logs = r.get_json()["data"]["logs"]
    assert isinstance(logs, list) and len(logs) > 0
    r2 = client.get(f"/api/jobs/{task_id}?since={len(logs)}", headers=headers)
    assert r2.status_code == 200
    assert r2.get_json()["data"]["logs"] == []
    r3 = client.get(f"/api/jobs/{task_id}?since={len(logs) + 5}", headers=headers)
    assert r3.status_code == 200
    assert r3.get_json()["data"]["logs"] == []


def test_jobs_list_contains_task(env, full_setup):
    """5. GET /api/jobs 列表包含该任务"""
    client, headers = env
    task_id = submit_job(client, headers, full_setup)
    wait_done(client, headers, task_id)
    r = client.get("/api/jobs", headers=headers)
    assert r.status_code == 200
    body = r.get_json()
    assert body["code"] == 0
    ids = [t["id"] for t in body["data"]]
    assert task_id in ids


def test_rerun_new_task_id(env, full_setup):
    """6. POST rerun → 新 task_id"""
    client, headers = env
    task_id = submit_job(client, headers, full_setup)
    wait_done(client, headers, task_id)
    r = client.post(f"/api/jobs/{task_id}/rerun", headers=headers)
    assert r.status_code == 200
    body = r.get_json()
    assert body["code"] == 0, body
    new_id = body["data"]["task_id"]
    assert new_id != task_id
    data = wait_done(client, headers, new_id)
    assert data["status"] == "done"


def test_delete_job(env, full_setup):
    """7. DELETE → 删除后 GET 404/code=4"""
    client, headers = env
    task_id = submit_job(client, headers, full_setup)
    wait_done(client, headers, task_id)
    r = client.delete(f"/api/jobs/{task_id}", headers=headers)
    assert r.status_code == 200
    assert r.get_json()["code"] == 0
    r2 = client.get(f"/api/jobs/{task_id}", headers=headers)
    assert r2.status_code == 404
    body = r2.get_json()
    assert body["code"] == 4, body
    # 重跑已删除任务 → 任务不存在
    r3 = client.post(f"/api/jobs/{task_id}/rerun", headers=headers)
    assert r3.status_code == 200
    assert r3.get_json()["code"] == 4


def test_tax_api_failure_still_done(env, full_setup, monkeypatch):
    """8. 税运接口失败 → 任务仍 done（标红行）不报错"""
    client, headers = env
    monkeypatch.setattr(engine, "get_charge_order_list", lambda order_no, cookie: FAKE_API_FAIL)
    monkeypatch.setattr(engine, "time", types.SimpleNamespace(sleep=lambda s: None))
    task_id = submit_job(client, headers, full_setup)
    data = wait_done(client, headers, task_id)
    assert data["status"] == "done", data
    assert data["error"] == ""
    assert data["result_files"]
    content = download_content(client, headers, task_id, data["result_files"][0])
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
    ws = wb["庆余处理表"]
    assert ws.max_row >= 3
    red = ws.cell(2, 1).fill.start_color.rgb
    assert red == "FFFFE5E5", f"失败行应标红，实际: {red}"
    wb.close()
