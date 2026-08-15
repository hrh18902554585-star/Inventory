# -*- coding: utf-8 -*-
"""测试 fixture 生成函数（从 smoke_test.py 抽出，供 pytest 复用）

- make_cost_xlsx / make_order_csv 接受文件路径或 file-like 对象
  （BytesIO / StringIO），便于通过 Flask test_client 直接构造 multipart 上传。
- 其余 make_* 接受路径参数，可写入 tmp_path。
- gen_fixtures 生成全套夹具目录（契约第 12 节，conftest.py 使用）。
"""
import csv
import os

import openpyxl

ORDER_CSV_HEADER = [
    "订单号", "订单状态", "商品id", "商家编码-规格维度",
    "商品数量(件)", "商家实收金额(元)", "平台优惠折扣(元)", "支付时间",
]

DEFAULT_COST_ROWS = [
    ["SKU-A", "", "雅培金装", 1, 50, 50],
    ["SKU-B", "", "雅培金装", 1, 60, 60],
]

_UNSET = object()


def make_cost_xlsx(target, headers=None, rows=None, baby_codes=_UNSET):
    """成本表：sku商品编码 sheet + 可选 母婴奶粉编码 sheet

    target 为路径或 file-like；默认含 SKU-A/SKU-B 两行及 CODE-999 奶粉编码；
    显式传 baby_codes=None 则不创建奶粉编码 sheet。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sku商品编码"
    ws.append(headers or ["组合编码", "单品编码", "商品名称", "数量", "单品成本", "sku成本"])
    for r in (rows if rows is not None else DEFAULT_COST_ROWS):
        ws.append(r)
    if baby_codes is _UNSET:
        baby_codes = ["CODE-999"]
    if baby_codes is not None:
        baby = wb.create_sheet("母婴奶粉编码")
        for code in baby_codes:
            baby.append([code])
    wb.save(target)
    return target


def make_order_csv(target, rows):
    """订单 CSV；target 为路径或 text-mode file-like（如 StringIO）"""
    if isinstance(target, (str, os.PathLike)):
        with open(target, "w", encoding="utf-8-sig", newline="") as f:
            _write_order_rows(f, rows)
    else:
        _write_order_rows(target, rows)
    return target


def _write_order_rows(f, rows):
    w = csv.writer(f)
    w.writerow(ORDER_CSV_HEADER)
    for r in rows:
        w.writerow(r)


def make_link_xlsx(path, headers=None, rows=None):
    """产品链接汇总表"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "产品链接汇总表"
    ws.append(headers or ["店铺名称", "产品名称", "商品id"])
    for r in (rows or []):
        ws.append(r)
    wb.save(path)
    return path


def make_promo_xlsx(path, rows=None, headers=None):
    """推广报表（文件名需含店铺名/日期供 extract_* 识别）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers if headers is not None else ["商品ID", "总花费(元)"])
    for r in (rows or []):
        ws.append(r)
    wb.save(path)
    return path


def make_subsidy_xlsx(path, headers=None, rows=None):
    """官补映射表"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers or ["链接id", "商家编码-规格维度", "官补金额"])
    for r in (rows or []):
        ws.append(r)
    wb.save(path)
    return path


def make_sort_xlsx(path, values=None):
    """产品排序表：单列，每行一个值"""
    wb = openpyxl.Workbook()
    ws = wb.active
    for v in (values or []):
        ws.append([v])
    wb.save(path)
    return path


def gen_fixtures(target_dir):
    """生成全套夹具到 target_dir，返回目录路径（str）"""
    os.makedirs(target_dir, exist_ok=True)
    make_cost_xlsx(os.path.join(target_dir, "成本表.xlsx"))
    make_link_xlsx(
        os.path.join(target_dir, "产品链接汇总表.xlsx"),
        rows=[
            ["庆余", "雅培金装", "1001"],
            ["趣味猴", "摩可纳8号", "2001"],
        ],
    )
    make_promo_xlsx(os.path.join(target_dir, "推广报表_庆余.xlsx"), rows=[["1001", 20]])
    make_promo_xlsx(os.path.join(target_dir, "推广报表_趣味猴.xlsx"), rows=[["2001", 30]])
    make_subsidy_xlsx(os.path.join(target_dir, "官补映射表.xlsx"), rows=[["1001", "SKU-A", 10]])
    make_sort_xlsx(os.path.join(target_dir, "产品汇总表.xlsx"), values=["雅培金装", "摩可纳8号"])
    make_order_csv(os.path.join(target_dir, "庆余订单_20260815.csv"), [
        ["20260001", "已发货，待收货", "1001", "SKU-A", 2, 200, 10, "2026-08-15 10:00:00"],
    ])
    return target_dir
