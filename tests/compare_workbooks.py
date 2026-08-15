# -*- coding: utf-8 -*-
"""Excel 黄金样本对比脚本
对比两个 xlsx（如 Web 引擎输出 vs GUI 版输出）的结构、值、样式一致性。

CLI: python compare_workbooks.py <a.xlsx> <b.xlsx> [--verbose]

对比项:
1. sheet 名称与顺序
2. 每个 sheet: 最大行列、合并单元格集合
3. 单元格值 (data_only=False 读公式文本):
   - 公式: 去除空白后按 "+" 拆段排序比较（容忍加数顺序/空白差异）
   - 非公式: 严格比较（含类型，None 与 "" 视为等价）
   - None 单元格跳过值对比（未写即空，等价）
4. 样式: fill (start_color/end_color)、border 有无、number_format、font bold
   （用 cell.has_style 快速跳过无样式单元格）
5. 列宽（容忍 ±0.5 舍入差，归入可容忍）

退出码: 0 = 无实质(致命)差异; 1 = 存在致命差异; 2 = 参数/文件错误
"""
import argparse
import os
import sys

import openpyxl


def setup_utf8():
    """Windows 控制台/管道下强制 UTF-8 输出（中文不乱码）"""
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


# ---------- 工具 ----------
def _empty(v):
    return v is None or (isinstance(v, str) and v == "")


def fmt_val(v):
    if v is None:
        return "(空)"
    if isinstance(v, str) and v == "":
        return '""'
    return repr(v)


def norm_formula(f):
    """公式规范化: 去空白、去前导 '='、按 '+' 拆段排序后重组"""
    s = "".join(str(f).split())
    if s.startswith("="):
        s = s[1:]
    return "+".join(sorted(s.split("+")))


def _fill_type(f):
    t = f.fill_type
    return "none" if t in (None, "none") else t


def _color_key(color):
    if color is None:
        return None
    rgb = getattr(color, "rgb", None)
    if rgb:
        return str(rgb).upper()
    if getattr(color, "theme", None) is not None:
        return "theme%d" % color.theme
    if getattr(color, "indexed", None) is not None:
        return "indexed%d" % color.indexed
    return None


def _border_key(b):
    # 边框只比"有无"（每边 style 是否设置）
    return tuple(s is not None for s in (b.left.style, b.right.style, b.top.style, b.bottom.style))


# ---------- 对比逻辑 ----------
def compare_value(va, vb):
    """单元格值对比。返回 (状态, 说明)，状态: skip/ok/tol/fatal"""
    if va is None and vb is None:
        return ("skip", "")
    ea, eb = _empty(va), _empty(vb)
    if ea and eb:
        return ("ok", "")
    if ea or eb:
        return ("fatal", "值: 期望=%s 实际=%s" % (fmt_val(va), fmt_val(vb)))
    fa = isinstance(va, str) and va.startswith("=")
    fb = isinstance(vb, str) and vb.startswith("=")
    if fa and fb:
        if va == vb:
            return ("ok", "")
        if norm_formula(va) == norm_formula(vb):
            return ("tol", "值: 公式加数顺序/空白差异(可容忍)")
        return ("fatal", "值: 期望=%s 实际=%s" % (fmt_val(va), fmt_val(vb)))
    if fa or fb:
        return ("fatal", "值: 公式性不一致 期望=%s 实际=%s" % (fmt_val(va), fmt_val(vb)))
    if isinstance(va, bool) != isinstance(vb, bool):
        return ("fatal", "值: 类型不一致 期望=%s 实际=%s" % (fmt_val(va), fmt_val(vb)))
    if isinstance(va, (int, float)) and isinstance(vb, (int, float)):
        if va == vb:
            return ("ok", "")
        return ("fatal", "值: 期望=%s 实际=%s" % (fmt_val(va), fmt_val(vb)))
    if type(va) is type(vb) and va == vb:
        return ("ok", "")
    return ("fatal", "值: 类型/值不一致 期望=%s 实际=%s" % (fmt_val(va), fmt_val(vb)))


def compare_style(ca, cb):
    """单元格样式对比（仅在任一侧 has_style 时调用）"""
    sa, sb = ca.has_style, cb.has_style
    if sa != sb:
        return ("fatal", "样式: has_style 不一致 期望=%s 实际=%s" % (sa, sb))
    if not sa:
        return ("ok", "")
    parts = []
    fa, fb = ca.fill, cb.fill
    ta, tb = _fill_type(fa), _fill_type(fb)
    if ta != tb:
        parts.append("fill类型 期望=%s 实际=%s" % (ta, tb))
    elif ta != "none":
        ka, kb = _color_key(fa.start_color), _color_key(fb.start_color)
        if ka != kb:
            parts.append("fill起始色 期望=%s 实际=%s" % (ka, kb))
        ka, kb = _color_key(fa.end_color), _color_key(fb.end_color)
        if ka != kb:
            parts.append("fill结束色 期望=%s 实际=%s" % (ka, kb))
    ba, bb = _border_key(ca.border), _border_key(cb.border)
    if ba != bb:
        parts.append("边框 期望=%s 实际=%s" % (ba, bb))
    if ca.number_format != cb.number_format:
        parts.append("数字格式 期望=%r 实际=%r" % (ca.number_format, cb.number_format))
    ba_, bb_ = bool(ca.font.bold), bool(cb.font.bold)
    if ba_ != bb_:
        parts.append("加粗 期望=%s 实际=%s" % (ba_, bb_))
    if parts:
        return ("fatal", "样式: " + "; ".join(parts))
    return ("ok", "")


def col_widths(ws):
    """提取单列显式宽度 {列字母: 宽度}"""
    d = {}
    for k, dim in ws.column_dimensions.items():
        if len(k) == 1 and k.isalpha() and dim.width:
            d[k] = float(dim.width)
    return d


def sheet_info(ws):
    return (ws.max_row, ws.max_column, sorted(str(r) for r in ws.merged_cells.ranges))


# ---------- 主对比 ----------
def compare_workbooks(path_a, path_b, verbose=False):
    wa = openpyxl.load_workbook(path_a, data_only=False)
    wb = openpyxl.load_workbook(path_b, data_only=False)

    diffs = []   # (sheet, coord, kind, msg)
    stat = {"fatal": 0, "tol": 0, "tol_formula": 0, "tol_width": 0}

    def add(sheet, coord, kind, msg):
        diffs.append((sheet, coord, kind, msg))
        stat["fatal" if kind == "致命" else "tol"] += 1

    na, nb = wa.sheetnames, wb.sheetnames
    if na != nb:
        add("*", "*", "致命", "结构: sheet 名称/顺序不一致 期望=%s 实际=%s" % (na, nb))
    if verbose:
        print("sheet 顺序: %s" % na)

    for sn in na:
        if sn not in nb:
            add(sn, "*", "致命", "结构: sheet 在 b 中缺失")
            continue
        wa_s, wb_s = wa[sn], wb[sn]
        info_a, info_b = sheet_info(wa_s), sheet_info(wb_s)
        if info_a[0] != info_b[0] or info_a[1] != info_b[1]:
            add(sn, "*", "致命", "结构: 最大行列不一致 期望=%dx%d 实际=%dx%d"
                % (info_a[0], info_a[1], info_b[0], info_b[1]))
        if info_a[2] != info_b[2]:
            add(sn, "*", "致命", "结构: 合并单元格集合不一致 期望=%s 实际=%s"
                % (info_a[2], info_b[2]))

        for col, (kw, msg) in compare_widths(wa_s, wb_s):
            stat["tol_width" if kw == "可容忍" else "fatal"] += 1
            add(sn, "%s列" % col, kw, msg)

        if verbose:
            print("\n== %s (期望 %dx%d, 实际 %dx%d, 合并 %s)" % (sn, info_a[0], info_a[1], info_b[0], info_b[1], info_a[2]))

        rows = max(info_a[0], info_b[0])
        cols = max(info_a[1], info_b[1])
        for r in range(1, rows + 1):
            for c in range(1, cols + 1):
                ca = wa_s.cell(r, c)
                cb = wb_s.cell(r, c)
                status, msg = compare_value(ca.value, cb.value)
                if status == "fatal":
                    add(sn, ca.coordinate, "致命", msg)
                elif status == "tol":
                    stat["tol_formula"] += 1
                    add(sn, ca.coordinate, "可容忍", msg)
                elif status == "ok" and verbose:
                    print("  OK  %s!%s = %s" % (sn, ca.coordinate, fmt_val(ca.value)))
                if ca.has_style or cb.has_style:
                    st, smsg = compare_style(ca, cb)
                    if st == "fatal":
                        add(sn, ca.coordinate, "致命", smsg)

    return diffs, stat


def compare_widths(wa_s, wb_s):
    da, db = col_widths(wa_s), col_widths(wb_s)
    out = []
    for col in sorted(set(da) | set(db)):
        a, b = da.get(col), db.get(col)
        if a is None or b is None:
            out.append((col, ("致命", "列宽: 期望=%s 实际=%s" % (a, b))))
        elif a == b:
            continue
        elif abs(a - b) <= 0.5:
            out.append((col, ("可容忍", "列宽: 舍入差 %.2f 期望=%.2f 实际=%.2f" % (abs(a - b), a, b))))
        else:
            out.append((col, ("致命", "列宽: 期望=%.2f 实际=%.2f" % (a, b))))
    return out


# ---------- 入口 ----------
def main():
    setup_utf8()
    parser = argparse.ArgumentParser(description="Excel 黄金样本对比（GUI vs 引擎输出）")
    parser.add_argument("a", help="期望文件 (如 GUI 输出)")
    parser.add_argument("b", help="实际文件 (如 Web 引擎输出)")
    parser.add_argument("--verbose", action="store_true", help="打印全部对比过程")
    args = parser.parse_args()

    for p in (args.a, args.b):
        if not os.path.isfile(p):
            print("错误: 文件不存在: %s" % p)
            return 2

    print("===== 对比: %s  vs  %s =====" % (args.a, args.b))
    diffs, stat = compare_workbooks(args.a, args.b, verbose=args.verbose)

    if args.verbose:
        print()
    for sheet, coord, kind, msg in diffs:
        print("[%s] %s!%s  %s" % (kind, sheet, coord, msg))

    total = len(diffs)
    print()
    print("===== 统计: 共 %d 处差异: 致命 %d 处, 可容忍 %d 处 (公式顺序 %d, 列宽 %d) ====="
          % (total, stat["fatal"], stat["tol"], stat["tol_formula"], stat["tol_width"]))
    if stat["fatal"]:
        print("结果: FAIL (存在致命差异)")
        return 1
    print("结果: PASS (无实质差异%s)" % ("，仅可容忍差异" if total else ""))
    return 0


if __name__ == "__main__":
    sys.exit(main())
