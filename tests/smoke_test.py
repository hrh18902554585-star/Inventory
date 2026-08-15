# -*- coding: utf-8 -*-
"""利润核算引擎冒烟测试
构造最小 fixture → 运行 ProfitEngine → 断言：
1. 引擎独立运行不抛异常
2. 生成的 Excel 结构（sheet 顺序/名称）正确
3. 处理表数值与手算一致
4. [R2 修复] 多店铺时销售计划登记表包含所有店铺数据
5. 空输入（无有效订单）时抛 EngineError
6. 多天数据生成合并文件
"""
import csv
import os
import shutil
import sys

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "profit_service"))

import openpyxl

from engine import EngineError, ProfitEngine

FIXTURES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures")
OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_smoke_out")

FAKE_COOKIE = "fake_cookie_for_smoke_test"
PASS = 0
FAIL = 0


def check(name, cond, detail=""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  [PASS] {name}")
    else:
        FAIL += 1
        print(f"  [FAIL] {name} {detail}")


# ---------- fixture 生成 ----------
def make_cost_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sku商品编码"
    ws.append(["组合编码", "单品编码", "商品名称", "数量", "单品成本", "sku成本"])
    ws.append(["SKU-A", "", "雅培金装", 1, 50, 50])
    ws.append(["SKU-B", "", "雅培金装", 1, 60, 60])
    baby = wb.create_sheet("母婴奶粉编码")
    baby.append(["CODE-999"])
    wb.save(path)


def make_link_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "产品链接汇总表"
    ws.append(["店铺名称", "产品名称", "商品id"])
    ws.append(["庆余", "雅培金装", "1001"])
    ws.append(["趣味猴", "摩可纳8号", "2001"])
    wb.save(path)


def make_promo_xlsx(path, store, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["商品ID", "总花费(元)"])
    for r in rows:
        ws.append(r)
    wb.save(path)


def make_subsidy_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["链接id", "规格编码", "官补金额"])
    ws.append(["1001", "SKU-A", 10])
    wb.save(path)


def make_sort_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["雅培金装"])
    ws.append(["摩可纳8号"])
    wb.save(path)


def make_order_csv(path, rows):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["订单号", "订单状态", "商品id", "商家编码-规格维度", "商品数量(件)", "商家实收金额(元)", "平台优惠折扣(元)", "支付时间"])
        for r in rows:
            w.writerow(r)


def gen_fixtures():
    shutil.rmtree(FIXTURES_DIR, ignore_errors=True)
    os.makedirs(FIXTURES_DIR)
    make_cost_xlsx(os.path.join(FIXTURES_DIR, "成本表.xlsx"))
    make_link_xlsx(os.path.join(FIXTURES_DIR, "产品链接汇总表.xlsx"))
    make_promo_xlsx(os.path.join(FIXTURES_DIR, "推广报表_庆余.xlsx"), "庆余", [["1001", 20]])
    make_promo_xlsx(os.path.join(FIXTURES_DIR, "推广报表_趣味猴.xlsx"), "趣味猴", [["2001", 30]])
    make_subsidy_xlsx(os.path.join(FIXTURES_DIR, "官补映射表.xlsx"))
    make_sort_xlsx(os.path.join(FIXTURES_DIR, "产品汇总表.xlsx"))
    make_order_csv(os.path.join(FIXTURES_DIR, "庆余订单_20260815.csv"), [
        ["20260001", "已发货，待收货", "1001", "SKU-A", 2, 200, 10, "2026-08-15 10:00:00"],
        ["20260002", "已收货", "1001", "SKU-B", 1, 150, 5, "2026-08-15 11:00:00"],
        ["20260003", "已退款", "1001", "SKU-A", 1, 100, 0, "2026-08-15 12:00:00"],
    ])
    make_order_csv(os.path.join(FIXTURES_DIR, "趣味猴订单_20260815.csv"), [
        ["20260004", "已收货", "2001", "SKU-C", 3, 300, 0, "2026-08-15 13:00:00"],
    ])
    make_order_csv(os.path.join(FIXTURES_DIR, "庆余订单_20260816.csv"), [
        ["20260005", "已收货", "1001", "SKU-A", 1, 100, 0, "2026-08-16 09:00:00"],
    ])
    make_order_csv(os.path.join(FIXTURES_DIR, "空订单.csv"), [])


def get_configs():
    return {
        "cost": os.path.join(FIXTURES_DIR, "成本表.xlsx"),
        "link": os.path.join(FIXTURES_DIR, "产品链接汇总表.xlsx"),
        "promo": [
            os.path.join(FIXTURES_DIR, "推广报表_庆余.xlsx"),
            os.path.join(FIXTURES_DIR, "推广报表_趣味猴.xlsx"),
        ],
        "subsidy": os.path.join(FIXTURES_DIR, "官补映射表.xlsx"),
        "sort": os.path.join(FIXTURES_DIR, "产品汇总表.xlsx"),
    }


# ---------- 测试 1: 单日单店铺 ----------
def test_single_day():
    print("\n== 测试1: 单日单店铺 ==")
    out = os.path.join(OUT_DIR, "single")
    shutil.rmtree(out, ignore_errors=True)
    logs = []

    engine = ProfitEngine(log_cb=logs.append, progress_cb=lambda p, t: None)
    result = engine.process(
        order_files=[os.path.join(FIXTURES_DIR, "庆余订单_20260815.csv")],
        configs=get_configs(),
        cookie=FAKE_COOKIE,
        thread_count=2,
        enable_tax=True,
        exclude_ids="",
        output_dir=out,
    )

    check("输出文件存在", len(result["files"]) == 1 and os.path.exists(result["files"][0]), result["files"])
    check("统计: 有效订单数=2 (退款行被过滤)", result["stats"]["total_tasks"] == 2, result["stats"])

    wb = openpyxl.load_workbook(result["files"][0], data_only=False)
    check("sheet 顺序正确", wb.sheetnames == ["庆余原表", "庆余处理表", "庆余推广表", "庆余明细表", "全店汇总表", "销售计划登记", "dokkan酵素香槟金每日数据", "movefree红瓶 200粒每日数据"], wb.sheetnames)

    ws = wb["庆余原表"]
    check("原表 4 行(表头+3数据含退款)", ws.max_row == 4, ws.max_row)

    ws = wb["庆余处理表"]
    # 追加列: 实际商品数量(n+1) 税运(n+2) 商品成本(n+3) 订单成本(n+4) 综合成本(n+5) 平台扣点(n+6) 毛利润(n+7) 归属日期(n+8) 费用项(n+9) + 费用名
    check("处理表 3 行(表头+2数据)", ws.max_row == 3, ws.max_row)
    check("行1 商品数量=2", ws.cell(2, 9).value == 2, ws.cell(2, 9).value)
    check("行1 税运=0(API失败)", ws.cell(2, 10).value == 0, ws.cell(2, 10).value)
    check("行1 商品成本=50", ws.cell(2, 11).value == 50, ws.cell(2, 11).value)
    check("行1 订单成本=100", ws.cell(2, 12).value == 100, ws.cell(2, 12).value)
    check("行1 综合成本=公式", str(ws.cell(2, 13).value).startswith("="), ws.cell(2, 13).value)
    check("行1 归属日期=2026年08月15日", ws.cell(2, 16).value == "2026年08月15日", ws.cell(2, 16).value)
    check("行2 订单成本=60", ws.cell(3, 12).value == 60, ws.cell(3, 12).value)

    ws = wb["全店汇总表"]
    prod_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == "雅培金装":
            prod_row = r
            break
    check("全店汇总含雅培金装", prod_row is not None)
    if prod_row:
        check("汇总 推广花费=公式(SUMIFS)", str(ws.cell(prod_row, 5).value).startswith("="), ws.cell(prod_row, 5).value)
        check("汇总 百亿官补=10", ws.cell(prod_row, 7).value == 10, ws.cell(prod_row, 7).value)

    ws = wb["销售计划登记"]
    check("销售计划 1 产品", ws.max_row == 2, ws.max_row)
    check("销售计划 产品名=雅培金装", ws.cell(2, 1).value == "雅培金装", ws.cell(2, 1).value)
    check("销售计划 销售额=公式", str(ws.cell(2, 2).value).startswith("="), ws.cell(2, 2).value)
    check("销售计划 推广花费=公式", str(ws.cell(2, 7).value).startswith("="), ws.cell(2, 7).value)
    wb.close()


# ---------- 测试 2: 多店铺 → R2 修复验证 ----------
def test_multi_store():
    print("\n== 测试2: 多店铺 (R2 修复验证) ==")
    out = os.path.join(OUT_DIR, "multi")
    shutil.rmtree(out, ignore_errors=True)

    engine = ProfitEngine()
    result = engine.process(
        order_files=[os.path.join(FIXTURES_DIR, "庆余订单_20260815.csv"), os.path.join(FIXTURES_DIR, "趣味猴订单_20260815.csv")],
        configs=get_configs(),
        cookie=FAKE_COOKIE,
        thread_count=2,
        enable_tax=True,
        exclude_ids="",
        output_dir=out,
    )

    wb = openpyxl.load_workbook(result["files"][0], data_only=False)
    ws = wb["销售计划登记"]
    check("销售计划 2 产品", ws.max_row == 3, ws.max_row)
    names = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
    check("销售计划 包含 雅培金装+摩可纳8号", "雅培金装" in names and "摩可纳8号" in names, names)
    mo_row = names.index("摩可纳8号") + 2
    check("摩可纳8号 销售额=公式(未被过滤)", str(ws.cell(mo_row, 2).value).startswith("="), ws.cell(mo_row, 2).value)

    ws2 = wb["庆余明细表"]
    mo_detail_row = None
    for r in range(2, ws2.max_row + 1):
        if ws2.cell(r, 2).value == "摩可纳8号":
            mo_detail_row = r
            break
    check("庆余明细表含摩可纳8号占位行", mo_detail_row is not None)
    if mo_detail_row:
        check("占位行 商品id 为空", ws2.cell(mo_detail_row, 3).value in (None, ""), ws2.cell(mo_detail_row, 3).value)
    wb.close()


# ---------- 测试 3: 空输入报错 ----------
def test_empty_input():
    print("\n== 测试3: 空输入报错 (R3a) ==")
    engine = ProfitEngine()
    try:
        engine.process(
            order_files=[os.path.join(FIXTURES_DIR, "空订单.csv")],
            configs=get_configs(),
            cookie=FAKE_COOKIE,
            enable_tax=False,
            output_dir=os.path.join(OUT_DIR, "empty"),
        )
        check("空输入抛 EngineError", False, "未抛出异常")
    except EngineError as e:
        check("空输入抛 EngineError", True, str(e))
    except Exception as e:
        check("空输入抛 EngineError", False, f"错误类型: {type(e)} {e}")


# ---------- 测试 4: 多天合并 ----------
def test_multi_day():
    print("\n== 测试4: 多天合并 ==")
    out = os.path.join(OUT_DIR, "multiday")
    shutil.rmtree(out, ignore_errors=True)

    engine = ProfitEngine()
    result = engine.process(
        order_files=[os.path.join(FIXTURES_DIR, "庆余订单_20260815.csv"), os.path.join(FIXTURES_DIR, "庆余订单_20260816.csv")],
        configs=get_configs(),
        cookie=FAKE_COOKIE,
        enable_tax=True,
        output_dir=out,
    )

    files = sorted(os.path.basename(f) for f in result["files"])
    check("生成 3 个文件(2单日+1合并)", len(result["files"]) == 3, files)

    merged = [f for f in files if "至" in f]
    check("合并文件存在", len(merged) == 1, merged)
    if merged:
        wb = openpyxl.load_workbook(os.path.join(out, merged[0]), data_only=False)
        ws = wb["庆余推广表"]
        check("合并推广表 3 列(含日期)", ws.max_column == 3, ws.max_column)
        ws = wb["庆余处理表"]
        check("合并处理表 4 行(2天共3单)", ws.max_row == 4, ws.max_row)
        wb.close()


# ---------- 主流程 ----------
def main():
    print("生成 fixtures ...")
    gen_fixtures()
    test_single_day()
    test_multi_store()
    test_empty_input()
    test_multi_day()
    print(f"\n===== 冒烟结果: {PASS} 通过, {FAIL} 失败 =====")
    sys.exit(1 if FAIL else 0)


if __name__ == "__main__":
    main()
