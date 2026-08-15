# -*- coding: utf-8 -*-
"""真实端到端冒烟（独立脚本，非 pytest）

模拟业务人员全流程，走真实 Flask app（create_app + test_client，无 monkeypatch 框架）：
1. 临时数据目录隔离（patch 各模块路径常量 → 不动 profit_service/data 共享数据）
2. 税运接口用 stub 替换 engine.get_charge_order_list（离线可跑，返回失败 → 任务标红但流程 done）
3. 上传 5 类配置 + 2 个订单 CSV → 建任务 → 轮询到 done → 下载校验 xlsx → 列表/统计

运行: venv\\Scripts\\python.exe tests\\e2e_smoke.py
退出码: 0 全 PASS，1 有 FAIL
"""
import io
import os
import re
import shutil
import sys
import tempfile
import time

if sys.stdout and hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

ROOT = os.path.dirname(os.path.abspath(__file__))
PROFIT_DIR = os.path.join(ROOT, "..", "profit_service")
for p in (PROFIT_DIR, ROOT):
    if p not in sys.path:
        sys.path.insert(0, p)

import openpyxl

from fixtures_gen import gen_fixtures, make_order_csv, make_subsidy_xlsx

PASS = 0
FAIL = 0


def check(name, cond, detail=""):
    global PASS, FAIL
    if cond:
        PASS += 1
        print(f"  [PASS] {name}")
    else:
        FAIL += 1
        print(f"  [FAIL] {name} {detail}")


def main():
    global PASS, FAIL
    os.makedirs(os.path.join(ROOT, "_e2e"), exist_ok=True)
    work = tempfile.mkdtemp(prefix="e2e_smoke_", dir=os.path.join(ROOT, "_e2e"))
    data_dir = os.path.join(work, "data")
    fix_dir = os.path.join(work, "fixtures")
    os.makedirs(data_dir, exist_ok=True)
    print(f"临时工作目录: {work}")

    # ---- 模块导入 + 路径常量 → 临时目录（必须在 create_app 之前） ----
    import logging_setup
    logging_setup.setup_logging(os.path.join(work, "logs"))  # 先初始化日志，避免写 profit_service/data/logs

    import app as app_mod
    import configs
    import db
    import engine
    import files_store
    import security
    import stats
    import task_manager

    configs.CONFIGS_DIR = os.path.join(data_dir, "configs")
    files_store.UPLOADS_DIR = os.path.join(data_dir, "uploads")
    security.TOKEN_FILE = os.path.join(data_dir, "token.txt")
    task_manager.TASKS_DIR = os.path.join(data_dir, "tasks")
    task_manager.OUTPUTS_DIR = os.path.join(data_dir, "outputs")
    db.DB_PATH = os.path.join(data_dir, "app.db")
    stats.STATS_FILE = os.path.join(data_dir, "stats.json")
    app_mod.DATA_DIR = data_dir
    app_mod.COOKIE_FILE = os.path.join(data_dir, "cookie.txt")

    # 税运接口 stub：返回失败，保证离线可跑（任务会标红但流程走通）
    engine.get_charge_order_list = lambda order_no, cookie: {"error": "e2e 离线 stub"}

    # ---- 真实启动 app ----
    print("\n== 1. 启动与鉴权 ==")
    app = app_mod.create_app()
    client = app.test_client()
    token = security.get_token()
    headers = {"X-Api-Token": token}
    check("create_app 启动成功（init_db/缓存/调度器/清扫）", app is not None)

    r = client.get("/")
    body = r.get_data(as_text=True)
    check("GET / 返回 200", r.status_code == 200, f"status={r.status_code}")
    check("首页含 vue.min.js 引用", "vue.min.js" in body)
    check("首页含 element-ui 引用", "element-ui" in body)

    r = client.get("/api/stats")
    check("无 token 请求返回 401", r.status_code == 401, f"status={r.status_code}")
    j = r.get_json()
    check("401 响应体 code=1 未授权", j and j.get("code") == 1 and j.get("message") == "未授权", str(j))

    r = client.get("/api/cookie", headers=headers)
    check("GET /api/cookie 200", r.status_code == 200, f"status={r.status_code}")
    j = r.get_json()
    check("初始 cookie_valid=False", j and j.get("data", {}).get("cookie_valid") is False, str(j))

    # ---- Cookie 保存 ----
    FAKE_COOKIE = "e2e_fake_cookie_12345"
    r = client.post("/api/cookie", json={"cookie": FAKE_COOKIE}, headers=headers)
    j = r.get_json()
    check("POST /api/cookie 保存成功", r.status_code == 200 and j and j.get("code") == 0, str(j))
    r = client.get("/api/cookie", headers=headers)
    j = r.get_json()
    d = j["data"]
    check("cookie_valid=True", d.get("cookie_valid") is True, str(d))
    check("掩码格式 ab***gh", re.fullmatch(r".{2}\*{3}.{2}", d.get("masked", "")) is not None,
          f"masked={d.get('masked')!r}")
    check("掩码不回传明文", FAKE_COOKIE not in d.get("masked", ""), str(d))

    # ---- 上传 5 类配置 ----
    print("\n== 2. 上传配置 ==")
    gen_fixtures(fix_dir)
    # fixtures_gen 默认官补表头为"商家编码-规格维度"，与 configs 上传校验要求"规格编码"不一致 → 显式修正（BUG#1）
    make_subsidy_xlsx(
        os.path.join(fix_dir, "官补映射表.xlsx"),
        headers=["链接id", "规格编码", "官补金额"],
        rows=[["1001", "SKU-A", 10]],
    )
    # 业务订单：庆余 3 行（2 有效 + 1 退款）+ 趣味猴 1 行
    make_order_csv(os.path.join(fix_dir, "庆余订单_20260815.csv"), [
        ["20260001", "已发货，待收货", "1001", "SKU-A", 2, 200, 10, "2026-08-15 10:00:00"],
        ["20260002", "已收货", "1001", "SKU-B", 1, 150, 5, "2026-08-15 11:00:00"],
        ["20260003", "已退款", "1001", "SKU-A", 1, 100, 0, "2026-08-15 12:00:00"],
    ])
    make_order_csv(os.path.join(fix_dir, "趣味猴订单_20260815.csv"), [
        ["20260004", "已收货", "2001", "SKU-C", 3, 300, 0, "2026-08-15 13:00:00"],
    ])

    def upload_config(cfg_type, filename):
        with open(os.path.join(fix_dir, filename), "rb") as f:
            raw = f.read()
        r = client.post("/api/configs/upload",
                        data={"type": cfg_type, "file": (io.BytesIO(raw), filename)},
                        headers=headers, content_type="multipart/form-data")
        j = r.get_json()
        ok = r.status_code == 200 and j and j.get("code") == 0
        detail = j if not ok else j["data"]
        check(f"上传配置 {cfg_type} ({filename})", ok, str(detail))
        return j["data"]["uploaded"][0] if ok else None

    cfg_ids = {}
    cfg_ids["cost"] = upload_config("cost", "成本表.xlsx")
    cfg_ids["link"] = upload_config("link", "产品链接汇总表.xlsx")
    cfg_ids["promo_qy"] = upload_config("promo", "推广报表_庆余.xlsx")
    cfg_ids["promo_qwh"] = upload_config("promo", "推广报表_趣味猴.xlsx")
    cfg_ids["subsidy"] = upload_config("subsidy", "官补映射表.xlsx")
    cfg_ids["sort"] = upload_config("sort", "产品汇总表.xlsx")

    r = client.get("/api/configs", headers=headers)
    j = r.get_json()
    data = j["data"]
    all_uploaded = all(v is not None for v in cfg_ids.values())
    cur = [data.get(t, {}).get("current", {}).get("version") for t in ("cost", "link", "subsidy", "sort")]
    check("GET /api/configs 5 类齐全且 current 就位", all_uploaded and all(cur), str(data))
    check("promo 上传 2 个文件", data.get("promo", {}).get("versions") and len(data["promo"]["versions"]) == 2,
          str(data.get("promo")))

    # ---- 上传 2 个订单 CSV ----
    print("\n== 3. 上传订单文件 ==")
    order_files = []
    for fn in ("庆余订单_20260815.csv", "趣味猴订单_20260815.csv"):
        with open(os.path.join(fix_dir, fn), "rb") as f:
            raw = f.read()
        r = client.post("/api/uploads",
                        data={"file": (io.BytesIO(raw), fn)},
                        headers=headers, content_type="multipart/form-data")
        j = r.get_json()
        ok = r.status_code == 200 and j and j.get("code") == 0
        detail = j if not ok else j["data"]
        check(f"上传订单 {fn}", ok, str(detail))
        if ok:
            order_files.append(j["data"]["uploaded"][0]["id"])
    check("订单文件列表 2 个", len(order_files) == 2, str(order_files))
    r = client.get("/api/uploads", headers=headers)
    j = r.get_json()
    check("GET /api/uploads 返回 2 条", j.get("code") == 0 and len(j["data"]) == 2, str(j))

    # ---- 建任务并轮询 ----
    print("\n== 4. 提交任务与轮询 ==")
    params = {
        "order_files": order_files,
        "configs": {
            "cost": cfg_ids["cost"]["version"],
            "link": cfg_ids["link"]["version"],
            "promo": [cfg_ids["promo_qy"]["version"], cfg_ids["promo_qwh"]["version"]],
            "subsidy": cfg_ids["subsidy"]["version"],
            "sort": cfg_ids["sort"]["version"],
        },
        "cookie": FAKE_COOKIE,
        "thread_count": 2,
        "enable_tax": True,
        "exclude_ids": "",
    }
    t0 = time.time()
    r = client.post("/api/jobs", json=params, headers=headers)
    j = r.get_json()
    check("POST /api/jobs 返回 task_id", r.status_code == 200 and j and j.get("code") == 0 and j["data"]["task_id"], str(j))
    task_id = j["data"]["task_id"] if j and j.get("code") == 0 else None

    if task_id:
        r = client.get(f"/api/jobs/{task_id}?since=0", headers=headers)
        j = r.get_json()
        st = j["data"]["status"]
        check("初始状态 pending/running", st in ("pending", "running"), st)

        deadline = time.time() + 30
        final = None
        while time.time() < deadline:
            r = client.get(f"/api/jobs/{task_id}?since=0", headers=headers)
            final = r.get_json()["data"]
            if final["status"] in ("done", "error"):
                break
            time.sleep(0.3)
        elapsed = time.time() - t0
        check("30s 内任务完成", final is not None and final["status"] in ("done", "error"), str(final))
        check("任务 status=done", final["status"] == "done", f"{final['status']} error={final.get('error')!r}")
        check("任务 error 为空", final.get("error") == "", repr(final.get("error")))
        check("任务 progress=100", final.get("progress") == 100, str(final.get("progress")))
        result_files = final.get("result_files") or []
        check("result_files 非空", len(result_files) > 0, str(result_files))
        logs = final.get("logs") or []
        print(f"  关键数据: 任务耗时 {elapsed:.1f}s | 日志条数 {len(logs)} | 输出文件 {result_files}")
        r = client.get(f"/api/jobs/{task_id}?since={len(logs)}", headers=headers)
        check("日志增量 since=N 返回空", r.get_json()["data"]["logs"] == [])

        # ---- 下载并校验 xlsx ----
        print("\n== 5. 下载与 xlsx 校验 ==")
        for fname in result_files:
            r = client.get(f"/api/jobs/{task_id}/download/{fname}", headers=headers)
            ok = r.status_code == 200 and r.data[:2] == b"PK"
            check(f"下载 {fname} 200 且为 xlsx", ok, f"status={r.status_code}")
            if ok:
                wb = openpyxl.load_workbook(io.BytesIO(r.data), data_only=False)
                need = ["庆余原表", "庆余处理表", "庆余推广表", "庆余明细表",
                        "趣味猴原表", "趣味猴处理表", "趣味猴推广表", "趣味猴明细表",
                        "全店汇总表", "销售计划登记"]
                missing = [s for s in need if s not in wb.sheetnames]
                check(f"{fname} sheet 齐全({len(need)}张)", not missing, f"缺失: {missing}")
                ws = wb["庆余处理表"]
                check("庆余处理表 表头+2行(退款行被过滤)", ws.max_row == 3, f"max_row={ws.max_row}")
                ws0 = wb["庆余原表"]
                check("庆余原表 表头+3行(含退款行)", ws0.max_row == 4, f"max_row={ws0.max_row}")
                check("行1 商品数量=2", ws.cell(2, 9).value == 2, str(ws.cell(2, 9).value))
                check("行1 税运=0(stub 失败标红)", ws.cell(2, 10).value == 0, str(ws.cell(2, 10).value))
                check("行1 标红填充", ws.cell(2, 1).fill.start_color.rgb == "FFFFE5E5",
                      str(ws.cell(2, 1).fill.start_color.rgb))
                wb.close()

        # ---- 列表 / 统计 ----
        print("\n== 6. 列表与统计 ==")
        r = client.get("/api/jobs", headers=headers)
        j = r.get_json()
        ids = [t["id"] for t in j["data"]]
        check("GET /api/jobs 列表含本任务", task_id in ids, str(ids))
        check("列表不泄露 cookie/日志", all("cookie" not in t and "logs" not in t for t in j["data"]), str(j["data"][0] if j["data"] else None))

        r = client.get("/api/stats", headers=headers)
        j = r.get_json()
        d = j["data"]
        check("GET /api/stats total_tasks>=1", d.get("total_tasks", 0) >= 1, str(d))
        check("GET /api/stats success>=1", d.get("success", 0) >= 1, str(d))
        print(f"  统计: {d}")

    # ---- 汇总 ----
    print(f"\n===== E2E 冒烟结果: {PASS} 通过, {FAIL} 失败 =====")
    if FAIL:
        print(f"临时目录保留供排查: {work}")
        sys.exit(1)
    shutil.rmtree(work, ignore_errors=True)
    sys.exit(0)


if __name__ == "__main__":
    main()
